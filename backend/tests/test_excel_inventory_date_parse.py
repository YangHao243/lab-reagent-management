"""Excel 库存宽表日期解析测试。

这些测试只验证本地解析函数，不连接线上数据库，也不读取 .env 中的生产连接。
"""

from __future__ import annotations

import os
import sys
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from sqlalchemy import select


BACKEND_DIR = Path(__file__).resolve().parents[1]
os.environ["DATABASE_URL"] = "sqlite:///file:excel_date_parse_pytest?mode=memory&cache=shared&uri=true"
sys.path.insert(0, str(BACKEND_DIR))

from services.excel_inventory_sync import (  # noqa: E402
    import_excel_inventory,
    parse_excel_row_date,
    parse_row_date,
    parse_sheet_year_month,
    row_has_operation,
)
from services.sync_core import ImportService, NormalizedInventoryRecord  # noqa: E402
from database import Base, SessionLocal, engine  # noqa: E402
import models  # noqa: E402,F401
from models import InventoryRecord, Reagent  # noqa: E402


def build_excel_content(
    row_number: int,
    day_value: object,
    operation: str | None,
    quantity: object,
    operator: object,
    sheet_name: str = "2026.5",
) -> bytes:
    """构造最小历史库存 Excel 宽表。"""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet["A2"] = "日期"
    worksheet["B2"] = "丙酮（MOS）"
    worksheet["B3"] = "操作"
    worksheet["C3"] = "数量"
    worksheet["D3"] = "操作人"
    worksheet.cell(row=row_number, column=1, value=day_value)
    worksheet.cell(row=row_number, column=2, value=operation)
    worksheet.cell(row=row_number, column=3, value=quantity)
    worksheet.cell(row=row_number, column=4, value=operator)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def reset_db() -> None:
    """重置测试库。"""

    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def fetch_inventory_records() -> list[InventoryRecord]:
    """读取全部库存流水，按 id 稳定排序。"""

    db = SessionLocal()
    try:
        return list(
            db.execute(select(InventoryRecord).order_by(InventoryRecord.id.asc()))
            .scalars()
            .all()
        )
    finally:
        db.close()


def make_normalized_record(
    *,
    operation_text: str = "领取",
    operation_type: str = "out",
    quantity: object = 20,
    operator: object = "张三",
    operation_time: datetime = datetime(2026, 5, 20, 10, 0, 0),
    source_row: int = 9,
    source_col: int = 2,
    remark: str | None = "Excel 导入：2026.5!R9C2",
) -> NormalizedInventoryRecord:
    """构造直接交给 ImportService 的标准库存流水。"""

    return NormalizedInventoryRecord(
        year=2026,
        month=5,
        event_date=operation_time.date(),
        reagent_name="丙酮（MOS）",
        operation_text=operation_text,
        operation_type=operation_type,
        quantity=float(quantity),
        operator=str(operator) if operator is not None else "",
        remark=remark,
        source="excel",
        source_sheet="2026.5",
        source_row=source_row,
        source_col=source_col,
        operation_time=operation_time,
    )


def test_parse_sheet_year_month_supports_common_separators() -> None:
    """sheet 名支持 2026.5、2026.05、2026-5、2026_5。"""

    assert parse_sheet_year_month("2026.5") == (2026, 5)
    assert parse_sheet_year_month("2026.05") == (2026, 5)
    assert parse_sheet_year_month("2026-5") == (2026, 5)
    assert parse_sheet_year_month("2026_5") == (2026, 5)


def test_parse_numeric_day_from_column_a() -> None:
    """sheet=2026.5 且 A9=6 时，应得到 2026-05-06。"""

    assert parse_row_date("2026.5", 9, 6) == date(2026, 5, 6)
    assert parse_row_date("2026.5", 9, 7.0) == date(2026, 5, 7)
    assert parse_excel_row_date("2026.5", 9, 6) == datetime(2026, 5, 6, 10, 0, 0)


def test_parse_string_day_from_column_a() -> None:
    """sheet=2026.5 且 A10='7'/'07' 时，应得到 2026-05-07。"""

    assert parse_row_date("2026.5", 10, "7") == date(2026, 5, 7)
    assert parse_row_date("2026.5", 10, "07") == date(2026, 5, 7)
    assert parse_excel_row_date("2026.5", 10, "07") == datetime(2026, 5, 7, 10, 0, 0)


def test_parse_first_day_of_april() -> None:
    """sheet=2026.4 且 A4=1 时，应得到 2026-04-01。"""

    assert parse_row_date("2026.4", 4, 1) == date(2026, 4, 1)


def test_parse_datetime_cell_must_match_sheet_month() -> None:
    """A 列是真实日期时使用日期本身，并校验 sheet 年月一致。"""

    assert parse_row_date("2026.5", 9, datetime(2026, 5, 6)) == date(2026, 5, 6)
    with pytest.raises(ValueError, match="不一致"):
        parse_row_date("2026.5", 9, datetime(2026, 4, 6))


def test_invalid_day_is_rejected() -> None:
    """sheet=2026.2 且 A 列 day=30 时应跳过，解析函数抛出明确错误。"""

    with pytest.raises(ValueError, match="日期超出当月范围"):
        parse_row_date("2026.2", 4, 30)


def test_blank_date_without_operation_can_be_skipped_without_warning() -> None:
    """A 列为空且该行没有任何操作记录时，可静默跳过。"""

    dataframe = pd.DataFrame(
        [
            [None, None, None, None],
            [None, "丙酮（MOS）", None, None],
            ["日期", "操作", "数量", "操作人"],
            [None, None, None, None],
        ]
    )

    assert row_has_operation(dataframe, 3, [("丙酮（MOS）", 1)]) is False


def test_blank_date_with_operation_is_rejected() -> None:
    """A 列为空但该行有操作记录时，应被识别并跳过导入。"""

    dataframe = pd.DataFrame(
        [
            [None, None, None, None],
            [None, "丙酮（MOS）", None, None],
            ["日期", "操作", "数量", "操作人"],
            [None, "入库", 40, "程浩航"],
        ]
    )

    assert row_has_operation(dataframe, 3, [("丙酮（MOS）", 1)]) is True
    with pytest.raises(ValueError, match="日期为空"):
        parse_row_date("2026.5", 9, None)


def test_import_excel_uses_column_a_date_at_10() -> None:
    """导入时应把 A 列日和 sheet 年月组合为当天 10:00:00。"""

    reset_db()
    db = SessionLocal()
    try:
        result = import_excel_inventory(
            db,
            "test.xlsx",
            build_excel_content(9, 6, "入库", 40, "程浩航"),
            operator_id=None,
        )
        db.commit()

        assert result.created == 1
        assert result.failed == 0
        record = db.execute(select(InventoryRecord)).scalar_one()
        assert record.created_at == datetime(2026, 5, 6, 10, 0, 0)
        assert record.event_date == date(2026, 5, 6)
        assert record.operator_name == "程浩航"
    finally:
        db.close()


def test_import_excel_blank_operator_defaults_to_dash_and_allows_negative_stock() -> None:
    """Excel 操作人为空时默认 '-'，领取导致负库存也允许导入。"""

    reset_db()
    db = SessionLocal()
    try:
        result = import_excel_inventory(
            db,
            "test.xlsx",
            build_excel_content(10, "07", "领取", 20, None),
            operator_id=None,
        )
        db.commit()

        assert result.created == 1
        assert result.failed == 0
        assert all("操作人不能为空" not in error.reason for error in result.errors)
        assert all("库存不足" not in error.reason for error in result.errors)

        record = db.execute(select(InventoryRecord)).scalar_one()
        assert record.created_at == datetime(2026, 5, 7, 10, 0, 0)
        assert record.operator_name == "-"
        assert record.quantity_change == -20
        assert record.before_quantity == 0
        assert record.after_quantity == -20

        reagent = db.get(Reagent, record.reagent_id)
        assert reagent is not None
        assert reagent.current_quantity == -20
    finally:
        db.close()


def test_import_excel_duplicate_record_is_skipped() -> None:
    """重复导入同一条历史记录时应跳过，不重复写库存流水。"""

    reset_db()
    db = SessionLocal()
    content = build_excel_content(9, 6, "入库", 40, "程浩航")
    try:
        first_result = import_excel_inventory(db, "test.xlsx", content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "test.xlsx", content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 0
        assert second_result.skipped == 1
        assert len(fetch_inventory_records()) == 1
    finally:
        db.close()


def test_import_excel_duplicate_out_record_uses_business_key_not_source_row() -> None:
    """同一业务流水即使来源行不同、领取/出库文案不同，也应跳过重复。"""

    reset_db()
    db = SessionLocal()
    first_content = build_excel_content(9, 20, "领取", 20, "张三")
    second_content = build_excel_content(10, 20, "出库", 20, "张三")
    try:
        first_result = import_excel_inventory(db, "first.xlsx", first_content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "second.xlsx", second_content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 0
        assert second_result.skipped == 1
        records = fetch_inventory_records()
        assert len(records) == 1
        assert records[0].created_at == datetime(2026, 5, 20, 10, 0, 0)
        assert records[0].operation_type == "out"
        assert records[0].quantity_change == -20
        assert records[0].operator_name == "张三"
    finally:
        db.close()


def test_import_excel_blank_operator_normalized_duplicate_is_skipped() -> None:
    """空操作人统一为 '-' 后，应能识别重复记录。"""

    reset_db()
    db = SessionLocal()
    first_content = build_excel_content(9, 20, "领取", 20, None)
    second_content = build_excel_content(10, 20, "领取", 20, "   ")
    try:
        first_result = import_excel_inventory(db, "first.xlsx", first_content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "second.xlsx", second_content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 0
        assert second_result.skipped == 1
        records = fetch_inventory_records()
        assert len(records) == 1
        assert records[0].operator_name == "-"
    finally:
        db.close()


def test_import_excel_signed_quantity_normalized_duplicate_is_skipped() -> None:
    """领取数量 20 和 -20 都会归一化为 -20，应识别为同一条。"""

    reset_db()
    db = SessionLocal()
    first_content = build_excel_content(9, 20, "领取", 20, "张三")
    second_content = build_excel_content(10, 20, "领取", -20, "张三")
    try:
        first_result = import_excel_inventory(db, "first.xlsx", first_content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "second.xlsx", second_content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 0
        assert second_result.skipped == 1
        assert len(fetch_inventory_records()) == 1
    finally:
        db.close()


def test_import_excel_different_date_is_not_duplicate() -> None:
    """日期不同即使其他字段相同，也应允许新增。"""

    reset_db()
    db = SessionLocal()
    first_content = build_excel_content(9, 20, "领取", 20, "张三")
    second_content = build_excel_content(10, 21, "领取", 20, "张三")
    try:
        first_result = import_excel_inventory(db, "first.xlsx", first_content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "second.xlsx", second_content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 1
        assert second_result.skipped == 0
        assert len(fetch_inventory_records()) == 2
    finally:
        db.close()


def test_import_excel_different_operator_is_not_duplicate() -> None:
    """操作员不同即使其他字段相同，也应允许新增。"""

    reset_db()
    db = SessionLocal()
    first_content = build_excel_content(9, 20, "领取", 20, "张三")
    second_content = build_excel_content(10, 20, "领取", 20, "李四")
    try:
        first_result = import_excel_inventory(db, "first.xlsx", first_content, operator_id=None)
        db.commit()
        second_result = import_excel_inventory(db, "second.xlsx", second_content, operator_id=None)
        db.commit()

        assert first_result.created == 1
        assert second_result.created == 1
        assert second_result.skipped == 0
        assert len(fetch_inventory_records()) == 2
    finally:
        db.close()


def test_import_excel_blank_date_with_operation_is_error_not_first_day() -> None:
    """A 列为空但该行有操作时跳过，并进入错误明细。"""

    reset_db()
    db = SessionLocal()
    try:
        result = import_excel_inventory(
            db,
            "test.xlsx",
            build_excel_content(9, None, "入库", 40, "程浩航"),
            operator_id=None,
        )
        db.commit()

        assert result.created == 0
        assert result.failed == 1
        assert "日期为空" in result.errors[0].reason
        assert fetch_inventory_records() == []
    finally:
        db.close()


def test_import_excel_invalid_day_is_error_not_crash() -> None:
    """非法日期只跳过该条记录，不让整个导入崩溃。"""

    reset_db()
    db = SessionLocal()
    try:
        result = import_excel_inventory(
            db,
            "test.xlsx",
            build_excel_content(4, 30, "入库", 40, "程浩航", sheet_name="2026.2"),
            operator_id=None,
        )
        db.commit()

        assert result.created == 0
        assert result.failed == 1
        assert "日期超出当月范围" in result.errors[0].reason
        assert fetch_inventory_records() == []
    finally:
        db.close()


def test_import_service_skips_duplicate_business_key_in_same_batch() -> None:
    """同一批次内业务字段相同但来源行列不同，只应新增一条。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(source_row=9, source_col=2, remark="Excel 导入：2026.5!R9C2"),
            make_normalized_record(source_row=10, source_col=5, remark="Excel 导入：2026.5!R10C5"),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 1
        assert result.failed == 0
        assert len(fetch_inventory_records()) == 1
    finally:
        db.close()


def test_import_service_business_dedup_ignores_remark_and_source_position() -> None:
    """remark/source_row/source_col 不参与业务去重。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(source_row=23, source_col=5, remark="Excel 导入：2026.5!R23C5"),
            make_normalized_record(source_row=24, source_col=8, remark="Excel 导入：2026.5!R24C8"),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 1
        assert len(fetch_inventory_records()) == 1
    finally:
        db.close()


def test_import_service_normalizes_out_synonyms_in_same_batch() -> None:
    """领取和出库归一化后应识别为同一条出库记录。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(operation_text="领取", operation_type="领取", source_row=9),
            make_normalized_record(operation_text="出库", operation_type="出库", source_row=10),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 1
        record = fetch_inventory_records()[0]
        assert record.operation_type == "out"
        assert record.quantity_change == -20
    finally:
        db.close()


def test_import_service_normalizes_signed_quantity_in_same_batch() -> None:
    """出库数量 20 和 -20 归一化后应识别为重复。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(operation_text="领取", operation_type="out", quantity=20, source_row=9),
            make_normalized_record(operation_text="领取", operation_type="out", quantity=-20, source_row=10),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 1
        assert len(fetch_inventory_records()) == 1
    finally:
        db.close()


def test_import_service_normalizes_blank_operator_in_same_batch() -> None:
    """空字符串、空格和 '-' 都归一化为 '-'，只新增一条。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(operator="", source_row=9),
            make_normalized_record(operator="   ", source_row=10),
            make_normalized_record(operator="-", source_row=11),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 2
        record = fetch_inventory_records()[0]
        assert record.operator_name == "-"
    finally:
        db.close()


def test_import_service_normalizes_microseconds_in_same_batch() -> None:
    """业务时间精确到秒，微秒不同不应导致重复导入。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(operation_time=datetime(2026, 5, 20, 10, 0, 0), source_row=9),
            make_normalized_record(operation_time=datetime(2026, 5, 20, 10, 0, 0, 123456), source_row=10),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 1
        assert result.skipped == 1
        assert fetch_inventory_records()[0].created_at == datetime(2026, 5, 20, 10, 0, 0)
    finally:
        db.close()


def test_import_service_different_date_or_operator_is_not_duplicate() -> None:
    """日期或操作员不同属于不同业务记录。"""

    reset_db()
    db = SessionLocal()
    try:
        records = [
            make_normalized_record(operator="张三", operation_time=datetime(2026, 5, 20, 10, 0, 0), source_row=9),
            make_normalized_record(operator="张三", operation_time=datetime(2026, 5, 21, 10, 0, 0), source_row=10),
            make_normalized_record(operator="李四", operation_time=datetime(2026, 5, 20, 10, 0, 0), source_row=11),
        ]
        result = ImportService(db).import_records(records)
        db.commit()

        assert result.created == 3
        assert result.skipped == 0
        assert len(fetch_inventory_records()) == 3
    finally:
        db.close()
