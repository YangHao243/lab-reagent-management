"""本地 Excel 库存表导入导出服务。

本模块只处理本地 Excel 宽表与标准化数据库之间的转换，不接入真实腾讯文档 API。
"""

from __future__ import annotations

import calendar
import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, select, text
from sqlalchemy.orm import Session

from database import engine
from models import InventoryRecord, Reagent
from services.sync_core import ImportService, NormalizedInventoryRecord, SyncImportResult
from utils.timezone import now_beijing


SHEET_NAME_PATTERN = re.compile(r"^\s*(\d{4})[._-](\d{1,2})\s*$")
OPERATION_MAPPING = {
    "入库": "in",
    "领取": "out",
    "出库": "out",
    "领用": "out",
}
OPERATION_LABELS = {
    "in": "入库",
    "out": "领取",
}
EXCEL_IMPORT_TIME = time(hour=10)


INVENTORY_RECORD_COLUMN_MIGRATIONS = {
    "event_date": "ALTER TABLE inventory_records ADD COLUMN event_date DATE",
    "source": "ALTER TABLE inventory_records ADD COLUMN source VARCHAR(50)",
    "source_sheet": "ALTER TABLE inventory_records ADD COLUMN source_sheet VARCHAR(100)",
    "source_row": "ALTER TABLE inventory_records ADD COLUMN source_row INTEGER",
    "source_col": "ALTER TABLE inventory_records ADD COLUMN source_col INTEGER",
    "source_hash": "ALTER TABLE inventory_records ADD COLUMN source_hash VARCHAR(64)",
    "updated_at": "ALTER TABLE inventory_records ADD COLUMN updated_at DATETIME",
}
SYNC_LOG_COLUMN_MIGRATIONS = {
    "detail_json": "ALTER TABLE sync_logs ADD COLUMN detail_json TEXT",
}


@dataclass
class ImportErrorItem:
    """单条导入错误。"""

    sheet: str | None
    row: int | None
    reagent: str | None
    reason: str

    def to_dict(self) -> dict[str, Any]:
        """转换为接口响应可序列化字典。"""

        return {
            "sheet": self.sheet,
            "row": self.row,
            "reagent": self.reagent,
            "reason": self.reason,
        }


@dataclass
class ExcelInventoryImportResult:
    """Excel 库存流水导入结果。"""

    created: int = 0
    skipped: int = 0
    failed: int = 0
    created_reagents: int = 0
    updated_reagents: int = 0
    errors: list[ImportErrorItem] = field(default_factory=list)

    @property
    def success(self) -> bool:
        """只要导入过程完成即视为成功，单行失败通过 failed/errors 表示。"""

        return True

    @property
    def message(self) -> str:
        """生成中文摘要。"""

        return (
            f"Excel 导入完成，新增 {self.created} 条，"
            f"跳过 {self.skipped} 条，失败 {self.failed} 条"
        )

    def add_error(
        self,
        sheet: str | None,
        row: int | None,
        reagent: str | None,
        reason: str,
    ) -> None:
        """记录一条错误并累计失败数。"""

        self.failed += 1
        self.errors.append(
            ImportErrorItem(sheet=sheet, row=row, reagent=reagent, reason=reason)
        )

    def to_detail_json(self) -> str:
        """保存到 SyncLog.detail_json 的结构化明细。"""

        return json.dumps(
            {
                "created": self.created,
                "skipped": self.skipped,
                "failed": self.failed,
                "created_reagents": self.created_reagents,
                "updated_reagents": self.updated_reagents,
                "errors": [error.to_dict() for error in self.errors[:200]],
            },
            ensure_ascii=False,
        )


def ensure_excel_sync_schema() -> None:
    """为旧版 SQLite 数据库幂等补齐 Excel 同步所需字段。"""

    inspector = inspect(engine)
    inventory_columns = {
        column["name"] for column in inspector.get_columns("inventory_records")
    }
    sync_log_columns = {
        column["name"] for column in inspector.get_columns("sync_logs")
    }

    with engine.begin() as connection:
        for column_name, migration_sql in INVENTORY_RECORD_COLUMN_MIGRATIONS.items():
            if column_name not in inventory_columns:
                connection.execute(text(migration_sql))

        for column_name, migration_sql in SYNC_LOG_COLUMN_MIGRATIONS.items():
            if column_name not in sync_log_columns:
                connection.execute(text(migration_sql))

        # SQLite 支持多个 NULL 通过唯一索引；导入记录 source_hash 非空时用于幂等去重。
        connection.execute(
            text(
                "CREATE UNIQUE INDEX IF NOT EXISTS "
                "ux_inventory_records_source_hash "
                "ON inventory_records(source_hash) "
                "WHERE source_hash IS NOT NULL"
            )
        )


def is_blank(value: Any) -> bool:
    """判断 Excel 单元格是否为空。"""

    return value is None or pd.isna(value) or str(value).strip() == ""


def clean_text(value: Any) -> str | None:
    """清理单元格文本。"""

    if is_blank(value):
        return None
    return str(value).strip()


def clean_quantity(value: Any) -> float:
    """解析数量，允许来源数据带正负号，但不能为 0。"""

    if is_blank(value):
        raise ValueError("数量不能为空")

    try:
        quantity = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("数量必须是数字") from exc

    if quantity == 0:
        raise ValueError("数量不能为 0")
    return quantity


def parse_sheet_year_month(sheet_name: str) -> tuple[int, int] | None:
    """从 sheet 名识别年份和月份，支持 2026.5 / 2026.05 / 2026-5 / 2026_5。"""

    match = SHEET_NAME_PATTERN.match(sheet_name)
    if match is None:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        return None
    return year, month


def parse_sheet_period(sheet_name: str) -> tuple[int, int] | None:
    """兼容旧函数名，内部统一使用 parse_sheet_year_month。"""

    return parse_sheet_year_month(sheet_name)


def row_has_operation(dataframe: pd.DataFrame, row_index: int, groups: list[tuple[str, int]]) -> bool:
    """判断当前数据行是否存在任意试剂操作内容。"""

    for _, column_index in groups:
        if column_index + 2 >= len(dataframe.columns):
            continue
        operation_text = clean_text(dataframe.iat[row_index, column_index])
        quantity_cell = dataframe.iat[row_index, column_index + 1]
        operator_name = clean_text(dataframe.iat[row_index, column_index + 2])
        if operation_text or not is_blank(quantity_cell) or operator_name:
            return True
    return False


def parse_row_date(sheet_name: str, row_number: int, row_date_cell: Any) -> date:
    """根据 sheet 年月和 A 列当前行日期生成业务日期。

    A 列为数字或数字字符串时只取“日”；A 列为真实日期/datetime 时优先使用
    该日期，同时校验它是否与 sheet 年月一致。
    """

    period = parse_sheet_year_month(sheet_name)
    if period is None:
        raise ValueError(f"无法从 sheet 名识别年月：{sheet_name}")

    year, month = period
    _ = row_number
    if is_blank(row_date_cell):
        raise ValueError("日期为空，已跳过")

    value = row_date_cell
    parsed_date: date | None = None

    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value) != int(value):
            raise ValueError(f"A列日期必须是整数日：{value}")
        day = int(value)
        max_day = calendar.monthrange(year, month)[1]
        if not 1 <= day <= max_day:
            raise ValueError(f"日期超出当月范围：{day}")
        return date(year, month, day)
    else:
        text_value = str(value).strip()
        if re.fullmatch(r"\d+(\.0+)?", text_value):
            day = int(float(text_value))
            max_day = calendar.monthrange(year, month)[1]
            if not 1 <= day <= max_day:
                raise ValueError(f"日期超出当月范围：{day}")
            return date(year, month, day)

        parsed_datetime = pd.to_datetime(text_value, errors="coerce")
        if pd.isna(parsed_datetime):
            raise ValueError(f"A列日期无法识别：{text_value}")
        parsed_date = parsed_datetime.date()

    if parsed_date.year != year or parsed_date.month != month:
        raise ValueError(
            f"A列日期 {parsed_date.isoformat()} 与 sheet 年月 {year}-{month:02d} 不一致"
        )
    return parsed_date


def parse_excel_row_date(sheet_name: str, row_number: int, row_date_cell: Any) -> datetime:
    """解析 Excel A 列日期，并统一补齐为当天 10:00:00。"""

    parsed_date = parse_row_date(sheet_name, row_number, row_date_cell)
    return datetime.combine(parsed_date, EXCEL_IMPORT_TIME)


def parse_day(value: Any, fallback_day: int, year: int, month: int) -> int:
    """兼容旧函数名：从 A 列解析日，空值不再按行号兜底。"""

    _ = fallback_day
    parsed_date = parse_row_date(f"{year}.{month}", 0, value)
    return parsed_date.day


def canonical_text(value: str | None) -> str:
    """用于试剂名称弱匹配的规范化文本。"""

    if not value:
        return ""
    return (
        value.lower()
        .replace(" ", "")
        .replace("\t", "")
        .replace("#", "")
        .replace("（", "(")
        .replace("）", ")")
    )


def parse_reagent_name(raw_name: str) -> tuple[str, str | None, str | None]:
    """从表头试剂名称提取标准名称、纯度和别名。"""

    name = raw_name.strip()
    parts = re.findall(r"[（(]([^）)]+)[）)]", name)
    base_name = re.sub(r"[（(][^）)]+[）)]", "", name).strip()
    purity_grade: str | None = None
    alias_parts: list[str] = []

    for part in parts:
        cleaned = part.strip()
        if cleaned.upper() in {"MOS", "AR", "GR", "CP"}:
            purity_grade = cleaned.upper()
        else:
            alias_parts.append(cleaned)

    standard_name = base_name or name
    alias_name = "；".join(alias_parts) if alias_parts else None
    return standard_name, purity_grade, alias_name


def format_reagent_label(reagent: Reagent) -> str:
    """导出时显示试剂名称，优先拼接纯度等级。"""

    if reagent.purity_grade:
        return f"{reagent.name_cn}（{reagent.purity_grade}）"
    return reagent.name_cn


def find_or_create_reagent(
    db: Session,
    raw_name: str,
    result: ExcelInventoryImportResult,
) -> Reagent:
    """按表头试剂名称查找或创建 Reagent，不覆盖库存字段。"""

    standard_name, purity_grade, alias_name = parse_reagent_name(raw_name)

    candidates = list(db.execute(select(Reagent)).scalars().all())
    raw_canonical = canonical_text(raw_name)
    standard_canonical = canonical_text(standard_name)
    alias_canonical = canonical_text(alias_name)

    for reagent in candidates:
        names = {
            canonical_text(reagent.name_cn),
            canonical_text(reagent.standard_name),
            canonical_text(reagent.alias_name),
        }
        if raw_canonical in names or standard_canonical in names or (
            alias_canonical and alias_canonical in names
        ):
            changed = False
            if not reagent.standard_name and standard_name:
                reagent.standard_name = standard_name
                changed = True
            if not reagent.purity_grade and purity_grade:
                reagent.purity_grade = purity_grade
                changed = True
            if not reagent.alias_name and alias_name:
                reagent.alias_name = alias_name
                changed = True
            if changed:
                result.updated_reagents += 1
            return reagent

    reagent = Reagent(
        name_cn=standard_name,
        standard_name=standard_name,
        purity_grade=purity_grade,
        alias_name=alias_name,
        unit="瓶",
        current_quantity=0.0,
        warning_threshold=1.0,
        is_preset=False,
    )
    db.add(reagent)
    db.flush()
    result.created_reagents += 1
    return reagent


def build_source_hash(
    sheet_name: str,
    excel_row: int,
    excel_col: int,
    reagent_name: str,
    operation_type: str,
    quantity: float,
    operator_name: str,
    event_date: date | datetime,
) -> str:
    """生成用于幂等导入的来源哈希。"""

    raw = "|".join(
        [
            "excel",
            sheet_name,
            str(excel_row),
            str(excel_col),
            reagent_name.strip(),
            operation_type,
            f"{quantity:g}",
            operator_name.strip(),
            event_date.isoformat(),
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def extract_reagent_groups(dataframe: pd.DataFrame) -> list[tuple[str, int]]:
    """从第 2 行提取试剂表头及其起始列。"""

    if len(dataframe.index) < 2:
        return []

    header_row = dataframe.iloc[1]
    groups: list[tuple[str, int]] = []
    for column_index, value in enumerate(header_row):
        if column_index == 0:
            continue

        reagent_name = clean_text(value)
        if not reagent_name:
            continue
        if reagent_name in {"操作", "数量", "操作人", "日期"}:
            continue

        groups.append((reagent_name, column_index))
    return groups


def import_excel_inventory(
    db: Session,
    file_name: str,
    content: bytes,
    operator_id: int | None,
) -> SyncImportResult:
    """导入历史 Excel 库存宽表，不提交事务。"""

    ensure_excel_sync_schema()
    parse_result = SyncImportResult()
    normalized_records: list[NormalizedInventoryRecord] = []
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xls":
        raise ValueError("当前环境暂不支持 .xls，请先另存为 .xlsx 后再导入")
    if suffix != ".xlsx":
        raise ValueError("历史库存流水导入仅支持 .xlsx 文件")

    excel_file = pd.ExcelFile(BytesIO(content), engine="openpyxl")
    seen_hashes: set[str] = set()

    for sheet_name in excel_file.sheet_names:
        period = parse_sheet_period(sheet_name)
        if period is None:
            continue

        year, month = period
        dataframe = pd.read_excel(
            excel_file,
            sheet_name=sheet_name,
            header=None,
            engine="openpyxl",
        )
        groups = extract_reagent_groups(dataframe)
        if not groups:
            parse_result.add_error(sheet_name, 2, None, "未识别到试剂名称表头")
            continue

        for row_index in range(3, min(len(dataframe.index), 34)):
            excel_row = row_index + 1
            try:
                operation_time = parse_excel_row_date(
                    sheet_name,
                    excel_row,
                    dataframe.iat[row_index, 0],
                )
                event_date = operation_time.date()
            except ValueError as exc:
                if row_has_operation(dataframe, row_index, groups):
                    affected_reagents: list[str] = []
                    for reagent_name, column_index in groups:
                        if column_index + 2 >= len(dataframe.columns):
                            continue
                        operation_text = clean_text(dataframe.iat[row_index, column_index])
                        quantity_cell = dataframe.iat[row_index, column_index + 1]
                        operator_name = clean_text(dataframe.iat[row_index, column_index + 2])
                        if operation_text or not is_blank(quantity_cell) or operator_name:
                            affected_reagents.append(reagent_name)
                    parse_result.add_error(
                        sheet_name,
                        excel_row,
                        "、".join(affected_reagents) if affected_reagents else None,
                        str(exc),
                    )
                continue

            for reagent_name, column_index in groups:
                if column_index + 2 >= len(dataframe.columns):
                    parse_result.add_error(
                        sheet_name,
                        excel_row,
                        reagent_name,
                        "试剂列不完整，缺少操作/数量/操作人中的字段",
                    )
                    continue

                operation_text = clean_text(dataframe.iat[row_index, column_index])
                quantity_cell = dataframe.iat[row_index, column_index + 1]
                operator_cell_text = clean_text(dataframe.iat[row_index, column_index + 2])

                if not operation_text and is_blank(quantity_cell) and not operator_cell_text:
                    continue

                if operation_text not in OPERATION_MAPPING:
                    parse_result.add_error(
                        sheet_name,
                        excel_row,
                        reagent_name,
                        "操作必须为入库、领取或出库",
                    )
                    continue

                try:
                    quantity = clean_quantity(quantity_cell)
                except ValueError as exc:
                    parse_result.add_error(sheet_name, excel_row, reagent_name, str(exc))
                    continue

                operator_name = operator_cell_text or "-"
                operation_type = OPERATION_MAPPING[operation_text]
                source_hash = build_source_hash(
                    sheet_name=sheet_name,
                    excel_row=excel_row,
                    excel_col=column_index + 1,
                    reagent_name=reagent_name,
                    operation_type=operation_type,
                    quantity=quantity,
                    operator_name=operator_name,
                    event_date=operation_time,
                )

                if source_hash in seen_hashes:
                    parse_result.skipped += 1
                    continue

                normalized_records.append(
                    NormalizedInventoryRecord(
                        year=year,
                        month=month,
                        event_date=event_date,
                        reagent_name=reagent_name,
                        operation_text=operation_text,
                        operation_type=operation_type,  # type: ignore[arg-type]
                        quantity=quantity,
                        operator=operator_name,
                        remark=f"Excel 导入：{sheet_name}!R{excel_row}C{column_index + 1}",
                        source="excel",
                        source_sheet=sheet_name,
                        source_row=excel_row,
                        source_col=column_index + 1,
                        source_hash=source_hash,
                        operation_time=operation_time,
                    )
                )
                seen_hashes.add(source_hash)

    import_result = ImportService(db=db, operator_id=operator_id).import_records(normalized_records)
    monthly_counts: dict[str, int] = {}
    for record in normalized_records:
        key = f"{record.year}-{record.month:02d}"
        monthly_counts[key] = monthly_counts.get(key, 0) + 1
    import_result.monthly_counts = monthly_counts
    import_result.skipped += parse_result.skipped
    import_result.failed += parse_result.failed
    import_result.errors = parse_result.errors + import_result.errors
    return import_result


def get_record_event_date(record: InventoryRecord) -> date:
    """返回库存流水业务日期，旧数据兜底使用 created_at。"""

    return record.event_date or record.created_at.date()


def aggregate_daily_records(
    records: list[InventoryRecord],
) -> dict[tuple[int, int, str], dict[str, Any]]:
    """按日、试剂、操作类型聚合库存流水。"""

    grouped: dict[tuple[int, int, str], dict[str, Any]] = {}
    for record in records:
        if record.operation_type not in {"in", "out"}:
            continue

        record_date = get_record_event_date(record)
        key = (record_date.month, record_date.day, record.reagent_id)
        item = grouped.setdefault(
            key,
            {
                "in_quantity": 0.0,
                "out_quantity": 0.0,
                "in_operators": set(),
                "out_operators": set(),
            },
        )

        operator_name = record.operator_name or ""
        if record.operation_type == "in":
            item["in_quantity"] += abs(record.quantity_change)
            if operator_name:
                item["in_operators"].add(operator_name)
        else:
            item["out_quantity"] += abs(record.quantity_change)
            if operator_name:
                item["out_operators"].add(operator_name)

    return grouped


def calculate_balance_before(
    records: list[InventoryRecord],
    reagent_id: int,
    before_date: date,
) -> float:
    """按库存流水计算指定日期前的结余。"""

    total = 0.0
    for record in records:
        if record.reagent_id != reagent_id:
            continue
        if get_record_event_date(record) < before_date:
            total += record.quantity_change
    return total


def calculate_balance_until(
    records: list[InventoryRecord],
    reagent_id: int,
    until_date: date,
) -> float:
    """按库存流水计算截至指定日期的结余。"""

    total = 0.0
    for record in records:
        if record.reagent_id != reagent_id:
            continue
        if get_record_event_date(record) <= until_date:
            total += record.quantity_change
    return total


def format_operators(operators: set[str]) -> str:
    """合并操作人显示。"""

    return "、".join(sorted(operators))


def export_inventory_excel(db: Session, year: int) -> Path:
    """按历史模板样式导出指定年份库存 Excel。"""

    ensure_excel_sync_schema()
    reagents = list(
        db.execute(
            select(Reagent).order_by(Reagent.display_order.asc(), Reagent.id.asc())
        ).scalars().all()
    )
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    records = list(
        db.execute(
            select(InventoryRecord).where(
                InventoryRecord.operation_type.in_(["in", "out"])
            )
        ).scalars().all()
    )
    records = [
        record
        for record in records
        if year_start <= get_record_event_date(record) <= year_end
    ]
    grouped_records = aggregate_daily_records(records)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sub_header_fill = PatternFill("solid", fgColor="EAF4E4")

    for month in range(1, 13):
        worksheet = workbook.create_sheet(f"{year}.{month}")
        worksheet.cell(row=1, column=1, value="上月结余")
        worksheet.cell(row=2, column=1, value="日期")
        worksheet.cell(row=3, column=1, value="日期")
        worksheet.column_dimensions["A"].width = 10

        for day in range(1, 32):
            worksheet.cell(row=day + 3, column=1, value=day)
        worksheet.cell(row=35, column=1, value="库存")

        for reagent_index, reagent in enumerate(reagents):
            start_col = 2 + reagent_index * 3
            end_col = start_col + 2
            label = format_reagent_label(reagent)

            worksheet.merge_cells(
                start_row=2,
                start_column=start_col,
                end_row=2,
                end_column=end_col,
            )
            worksheet.cell(row=2, column=start_col, value=label)
            worksheet.cell(row=3, column=start_col, value="操作")
            worksheet.cell(row=3, column=start_col + 1, value="数量")
            worksheet.cell(row=3, column=start_col + 2, value="操作人")

            month_start = date(year, month, 1)
            month_end = date(year, month, calendar.monthrange(year, month)[1])
            worksheet.cell(
                row=1,
                column=start_col,
                value=calculate_balance_before(records, reagent.id, month_start),
            )
            worksheet.cell(
                row=35,
                column=start_col,
                value=calculate_balance_until(records, reagent.id, month_end),
            )

            for day in range(1, calendar.monthrange(year, month)[1] + 1):
                row_number = day + 3
                item = grouped_records.get((month, day, reagent.id))
                if item is None:
                    continue

                operation_labels: list[str] = []
                quantity_labels: list[str] = []
                operators: set[str] = set()
                if item["in_quantity"] > 0:
                    operation_labels.append("入库")
                    quantity_labels.append(f"入库 {item['in_quantity']:g}")
                    operators.update(item["in_operators"])
                if item["out_quantity"] > 0:
                    operation_labels.append("领取")
                    quantity_labels.append(f"领取 {item['out_quantity']:g}")
                    operators.update(item["out_operators"])

                worksheet.cell(row=row_number, column=start_col, value="/".join(operation_labels))
                worksheet.cell(row=row_number, column=start_col + 1, value="；".join(quantity_labels))
                worksheet.cell(row=row_number, column=start_col + 2, value=format_operators(operators))

            for column in range(start_col, end_col + 1):
                worksheet.column_dimensions[get_column_letter(column)].width = 14
                worksheet.cell(row=2, column=column).fill = header_fill
                worksheet.cell(row=3, column=column).fill = sub_header_fill
                worksheet.cell(row=2, column=column).alignment = Alignment(horizontal="center")
                worksheet.cell(row=3, column=column).alignment = Alignment(horizontal="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row in {1, 2, 3, 35}:
                    cell.font = Font(bold=True)

    export_dir = Path(__file__).resolve().parent.parent / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = now_beijing().strftime("%Y%m%d_%H%M%S")
    file_path = export_dir / f"excel_inventory_{year}_{timestamp}.xlsx"
    workbook.save(file_path)
    return file_path
