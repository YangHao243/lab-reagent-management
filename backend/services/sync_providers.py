"""同步 Provider 抽象与实现。"""

from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode, urljoin

import requests
from sqlalchemy import select
from sqlalchemy.orm import Session

from config import settings
from models import SyncLog, TencentDocsToken
from services.excel_inventory_sync import (
    OPERATION_MAPPING,
    clean_quantity,
    clean_text,
    export_inventory_excel,
    import_excel_inventory,
    is_blank,
    parse_excel_row_date,
    parse_sheet_year_month,
)
from services.sync_core import (
    ExportService,
    ImportService,
    NormalizedInventoryRecord,
    SyncImportResult,
    SyncLogService,
)
from services.tencent_docs_matrix import (
    MATRIX_READ_RANGE,
    MATRIX_WRITE_RANGE,
    MATRIX_RANGE,
    MATRIX_TEMPLATE_TYPE,
    WRITE_COL_COUNT,
    WRITE_ROW_COUNT,
    build_reagent_matrix_patches_from_db,
    merge_patch_with_existing,
    parse_reagent_matrix,
    validate_patch_values,
)
from services.token_expiry import get_token_expiry_info
from utils.timezone import now_beijing


def get_token_file_path() -> Path:
    """返回本地腾讯文档 token 文件路径。"""

    token_file = Path(settings.TENCENT_DOC_TOKEN_FILE)
    if not token_file.is_absolute():
        token_file = Path(__file__).resolve().parent.parent / token_file
    return token_file


class TencentDocsConfigError(RuntimeError):
    """腾讯文档真实同步配置错误。"""


class TencentDocsApiNotImplementedError(NotImplementedError):
    """腾讯文档真实 OpenAPI 尚未实现。"""


class TencentDocsApiError(RuntimeError):
    """腾讯文档真实 OpenAPI 调用错误。"""

    def __init__(self, message: str, detail: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.detail = detail or {}


class TencentDocsEndpointConfigError(RuntimeError):
    """腾讯文档表格读写 API endpoint 未配置。"""

    def __init__(self, message: str, diagnosis: dict[str, Any]) -> None:
        super().__init__(message)
        self.diagnosis = diagnosis


PLACEHOLDER_CONFIG_VALUES = {
    "",
    "your-web-url-file-id",
    "your-sheetbook-book-id",
    "your-tab-id",
    "your-sheet-id",
    "your-sheet-title",
    "your-doc-file-id",
}


def clean_config_value(value: Any) -> str:
    """清理配置值，避免把 .env.example 中的占位符当成有效配置。"""

    text = str(value or "").strip()
    return "" if text.lower() in PLACEHOLDER_CONFIG_VALUES else text


def looks_like_docs_encoded_id(value: str) -> bool:
    """判断是否更像 docs.qq.com/sheet/{encodedID} 中的网页 encoded id。

    腾讯文档网页 slug 常见形态是 D 开头的短字符串，例如 DRmxPc2Fob2pFb252。
    这类值不能直接当成 OpenAPI fileID/bookID 使用，需要先走 converter。
    """

    text = value.strip()
    return bool(text) and text.startswith("D") and "/" not in text and "?" not in text


def get_tencent_docs_config() -> dict[str, Any]:
    """读取真实腾讯文档同步配置，TENCENT_DOCS_* 优先，旧字段兜底。"""

    raw_mode = (settings.TENCENT_DOCS_MODE or "").strip().lower()
    mode = raw_mode if raw_mode in {"real", "mock", "local"} else "local"
    raw_file_id = clean_config_value(settings.TENCENT_DOCS_FILE_ID) or clean_config_value(
        settings.TENCENT_DOC_DOC_ID
    )
    encoded_id = clean_config_value(settings.TENCENT_DOCS_ENCODED_ID)
    file_id = raw_file_id
    if raw_file_id and looks_like_docs_encoded_id(raw_file_id):
        encoded_id = encoded_id or raw_file_id
        file_id = ""

    sheet_id = clean_config_value(settings.TENCENT_DOCS_SHEET_ID)
    legacy_tab_id = clean_config_value(settings.TENCENT_DOCS_TAB_ID) or clean_config_value(
        settings.TENCENT_DOC_SHEET_ID
    )
    return {
        "mode": mode,
        "client_id": clean_config_value(settings.TENCENT_DOCS_CLIENT_ID)
        or clean_config_value(settings.TENCENT_DOC_CLIENT_ID),
        "client_secret": clean_config_value(settings.TENCENT_DOCS_CLIENT_SECRET)
        or clean_config_value(settings.TENCENT_DOC_CLIENT_SECRET),
        "redirect_uri": clean_config_value(settings.TENCENT_DOCS_REDIRECT_URI)
        or clean_config_value(settings.TENCENT_DOC_REDIRECT_URI),
        "access_token": clean_config_value(settings.TENCENT_DOCS_ACCESS_TOKEN),
        "open_id": clean_config_value(settings.TENCENT_DOCS_OPEN_ID),
        "encoded_id": encoded_id,
        "raw_file_id": raw_file_id,
        "file_id": file_id,
        "book_id": clean_config_value(settings.TENCENT_DOCS_BOOK_ID),
        "sheet_id": sheet_id,
        "sheet_title": clean_config_value(settings.TENCENT_DOCS_SHEET_TITLE),
        "tab_id": legacy_tab_id,
        "doc_id": file_id or encoded_id or clean_config_value(settings.TENCENT_DOC_DOC_ID),
        "api_base_url": clean_config_value(settings.TENCENT_DOCS_API_BASE_URL) or "https://docs.qq.com",
        # 读写路径已由后端按官方 OpenAPI 固定构造，保留字段仅用于兼容旧配置展示。
        "read_path": "",
        "update_path": "",
        "sheet_range": clean_config_value(settings.TENCENT_DOCS_SHEET_RANGE)
        or clean_config_value(settings.TENCENT_DOC_SHEET_RANGE)
        or MATRIX_RANGE,
        "read_range": clean_config_value(settings.TENCENT_DOCS_READ_RANGE) or MATRIX_READ_RANGE,
        "write_range": clean_config_value(settings.TENCENT_DOCS_WRITE_RANGE) or MATRIX_WRITE_RANGE,
        "template_type": clean_config_value(settings.TENCENT_DOCS_TEMPLATE_TYPE) or MATRIX_TEMPLATE_TYPE,
        "sheet_map_json": (settings.TENCENT_DOCS_SHEET_MAP_JSON or "").strip(),
        "sheet_title_pattern": (settings.TENCENT_DOCS_SHEET_TITLE_PATTERN or "").strip() or "{year}.{month}",
        "active_month": int(settings.TENCENT_DOCS_ACTIVE_MONTH or 1),
        "default_year": settings.TENCENT_DOCS_DEFAULT_YEAR,
        "write_cell_test_sheet_id": clean_config_value(settings.TENCENT_DOCS_WRITE_CELL_TEST_SHEET_ID) or "000001",
        "write_cell_test_range": clean_config_value(settings.TENCENT_DOCS_WRITE_CELL_TEST_RANGE) or "A1:A1",
    }


def build_sync_status() -> dict[str, Any]:
    """构建统一同步状态。"""

    real_status = RealTencentDocsProvider().get_status()
    has_client_id = bool(real_status["client_id_configured"])
    has_client_secret = bool(real_status["client_secret_configured"])
    has_redirect_uri = bool(real_status["redirect_uri_configured"])
    has_access_token = bool(real_status["access_token_configured"])
    has_doc_id = bool(real_status["doc_id_configured"])
    token_saved = bool(real_status["token_saved"])
    tencent_docs_enabled = bool(
        real_status.get("ready_for_sync")
        or real_status.get("ready_for_import")
        or real_status.get("ready_for_export")
    )
    mode = real_status["mode"] if real_status["mode"] == "real" else "local"

    return {
        "mode": mode,
        "mock_enabled": settings.SYNC_MOCK_ENABLED,
        "excel_enabled": True,
        "tencent_docs_enabled": tencent_docs_enabled,
        "client_id_configured": has_client_id,
        "client_secret_configured": has_client_secret,
        "redirect_uri_configured": has_redirect_uri,
        "access_token_configured": has_access_token,
        "doc_id_configured": has_doc_id,
        "file_id_configured": bool(real_status.get("file_id_configured")),
        "encoded_id_configured": bool(real_status.get("encoded_id_configured")),
        "file_id_resolved": bool(real_status.get("file_id_resolved")),
        "book_id_configured": bool(real_status.get("book_id_configured")),
        "book_id_resolved": bool(real_status.get("book_id_resolved")),
        "sheet_id_configured": bool(real_status.get("sheet_id_configured")),
        "sheet_id_resolved": bool(real_status.get("sheet_id_resolved")),
        "sheet_range_configured": bool(real_status.get("sheet_range_configured")),
        "tab_id_configured": bool(real_status.get("tab_id_configured")),
        "token_saved": token_saved,
        # 兼容旧前端字段。
        "has_client_id": has_client_id,
        "has_client_secret": has_client_secret,
        "has_redirect_uri": has_redirect_uri,
        "has_access_token": has_access_token,
        "has_doc_id": has_doc_id,
        "has_sheet_id": bool(settings.TENCENT_DOC_SHEET_ID.strip()),
        "has_token": token_saved,
        "doc_id": real_status["doc_id"],
        "file_id": real_status.get("file_id"),
        "book_id": real_status.get("book_id"),
        "sheet_id": real_status.get("sheet_id"),
        "sheet_title": real_status.get("sheet_title"),
        "tab_id": real_status.get("tab_id"),
        "token_valid": real_status["token_valid"],
        "token_expires_at": real_status["token_expires_at"],
        "open_id_saved": real_status["open_id_saved"],
        "auth_mode": real_status["auth_mode"],
        "ready_for_oauth": real_status["ready_for_oauth"],
        "ready_for_direct_token": real_status["ready_for_direct_token"],
        "sheet_read_endpoint_configured": real_status["sheet_read_endpoint_configured"],
        "sheet_write_endpoint_configured": real_status["sheet_write_endpoint_configured"],
        "ready_for_api_endpoint": real_status["ready_for_api_endpoint"],
        "ready_for_import": bool(real_status.get("ready_for_import")),
        "ready_for_export": bool(real_status.get("ready_for_export")),
        "ready_for_sync": real_status["ready_for_sync"],
        "last_probe_error": real_status.get("last_probe_error"),
        "description": (
            "当前为真实腾讯文档 API 模式"
            if tencent_docs_enabled
            else "当前为模拟/本地同步模式，可测试导入导出流程；不会访问真实腾讯文档。"
        ),
    }


class SyncProvider(ABC):
    """同步 Provider 基类。"""

    source: str

    @abstractmethod
    def get_status(self) -> dict[str, Any]:
        """返回 Provider 状态。"""

    @abstractmethod
    def validate_config(self) -> None:
        """校验 Provider 配置。"""

    @abstractmethod
    def import_records(self, db: Session, operator_id: int | None = None, **kwargs: Any) -> SyncImportResult:
        """导入标准化库存流水。"""

    @abstractmethod
    def export_records(self, db: Session, year: int = 2026, **kwargs: Any) -> Any:
        """导出数据。"""


class ExcelSyncProvider(SyncProvider):
    """Excel/CSV 本地同步 Provider。"""

    source = "excel"

    def get_status(self) -> dict[str, Any]:
        return {"excel_enabled": True}

    def validate_config(self) -> None:
        return None

    def import_records(
        self,
        db: Session,
        operator_id: int | None = None,
        **kwargs: Any,
    ) -> SyncImportResult:
        file_name = str(kwargs["file_name"])
        content = kwargs["content"]
        return import_excel_inventory(
            db=db,
            file_name=file_name,
            content=content,
            operator_id=operator_id,
        )

    def export_records(self, db: Session, year: int = 2026, **kwargs: Any) -> Path:
        _ = kwargs
        return export_inventory_excel(db, year=year)


class MockTencentDocsProvider(SyncProvider):
    """模拟腾讯文档 Provider，用于本地联调同步流程。"""

    source = "tencent_docs_mock"

    def get_status(self) -> dict[str, Any]:
        return {"mock_enabled": settings.SYNC_MOCK_ENABLED}

    def validate_config(self) -> None:
        if not settings.SYNC_MOCK_ENABLED:
            raise RuntimeError("Mock 同步已关闭")

    def import_records(
        self,
        db: Session,
        operator_id: int | None = None,
        **kwargs: Any,
    ) -> SyncImportResult:
        _ = kwargs
        self.validate_config()
        records = self.build_mock_records()
        return ImportService(db=db, operator_id=operator_id).import_records(records)

    def export_records(self, db: Session, year: int = 2026, **kwargs: Any) -> list[dict[str, Any]]:
        _ = kwargs
        self.validate_config()
        return ExportService(db=db).export_rows(year=year)

    def build_mock_records(self) -> list[NormalizedInventoryRecord]:
        """构造接近真实腾讯文档结构的 mock 库存流水。"""

        return [
            NormalizedInventoryRecord(
                year=2026,
                month=1,
                event_date=date(2026, 1, 5),
                reagent_name="丙酮（MOS）",
                operation_text="入库",
                operation_type="in",
                quantity=5,
                operator="Mock张三",
                remark="Mock 腾讯文档导入",
                source=self.source,
                source_sheet="2026.1",
                source_row=4,
                source_col=2,
            ),
            NormalizedInventoryRecord(
                year=2026,
                month=1,
                event_date=date(2026, 1, 6),
                reagent_name="丙酮（MOS）",
                operation_text="领取",
                operation_type="out",
                quantity=1,
                operator="Mock李四",
                remark="Mock 腾讯文档导入",
                source=self.source,
                source_sheet="2026.1",
                source_row=5,
                source_col=2,
            ),
            NormalizedInventoryRecord(
                year=2026,
                month=1,
                event_date=date(2026, 1, 7),
                reagent_name="无水乙醇（MOS）",
                operation_text="入库",
                operation_type="in",
                quantity=3,
                operator="Mock王五",
                remark="Mock 腾讯文档导入",
                source=self.source,
                source_sheet="2026.1",
                source_row=6,
                source_col=5,
            ),
        ]


class RealTencentDocsProvider(SyncProvider):
    """真实腾讯文档 Provider 骨架。

    本阶段只负责配置读取、状态判断和接口占位，不发起真实网络请求。
    """

    source = "tencent_docs_real"

    def __init__(self, db: Session | None = None) -> None:
        self.db = db
        self._resolved_file_id: str | None = None
        self._resolved_book_id: str | None = None
        self._resolved_sheets: list[dict[str, Any]] | None = None
        self._resolved_sheet: dict[str, Any] | None = None
        self._last_probe_error: str | None = None

    def _latest_token(self) -> TencentDocsToken | None:
        """读取最近一条腾讯文档 token。无数据库会话时返回 None。"""

        if self.db is None:
            return None

        stmt = (
            select(TencentDocsToken)
            .where(TencentDocsToken.provider == "tencent_docs")
            .order_by(TencentDocsToken.id.desc())
            .limit(1)
        )
        return self.db.execute(stmt).scalar_one_or_none()

    def _missing_oauth_fields(self) -> list[str]:
        config = get_tencent_docs_config()
        missing_fields: list[str] = []
        if not config["client_id"]:
            missing_fields.append("TENCENT_DOCS_CLIENT_ID")
        if not config["client_secret"]:
            missing_fields.append("TENCENT_DOCS_CLIENT_SECRET")
        if not config["redirect_uri"]:
            missing_fields.append("TENCENT_DOCS_REDIRECT_URI")
        return missing_fields

    def _missing_direct_token_fields(self) -> list[str]:
        config = get_tencent_docs_config()
        missing_fields: list[str] = []
        if not config["client_id"]:
            missing_fields.append("TENCENT_DOCS_CLIENT_ID")
        if not config["access_token"]:
            missing_fields.append("TENCENT_DOCS_ACCESS_TOKEN")
        if not config["open_id"]:
            missing_fields.append("TENCENT_DOCS_OPEN_ID")
        return missing_fields

    def _missing_sheetbook_fields(self) -> list[str]:
        config = get_tencent_docs_config()
        missing_fields: list[str] = []
        if not config["api_base_url"]:
            missing_fields.append("TENCENT_DOCS_API_BASE_URL")
        if not config["book_id"]:
            missing_fields.append("TENCENT_DOCS_BOOK_ID")
        return missing_fields

    def build_endpoint_diagnosis(
        self,
        *,
        operation: str,
        attempted_endpoint: str | None = None,
        attempts: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        """构造 endpoint 缺失诊断，不包含 token 明文。"""

        config = get_tencent_docs_config()
        return {
            "direct_token_configured": bool(
                config["client_id"] and config["access_token"] and config["open_id"]
            ),
            "encoded_id_configured": bool(config["encoded_id"]),
            "file_id_configured": bool(config["file_id"]),
            "book_id_configured": bool(config["book_id"]),
            "sheet_id_configured": bool(config["sheet_id"]),
            "sheet_title_configured": bool(config["sheet_title"]),
            "tab_id_configured": bool(config["tab_id"]),
            "api_base_url_configured": bool(config["api_base_url"]),
            "sheet_range_configured": bool(config["sheet_range"]),
            "sheetbook_api_base": "/openapi/sheetbook/v2/{bookID}",
            "spreadsheet_read_api": "/openapi/spreadsheet/v3/files/{fileId}/{sheetId}/{range}",
            "attempted_endpoint": attempted_endpoint,
            "operation": operation,
            "conversion_attempts": attempts or [],
            "next_step": (
                "Configure TENCENT_DOCS_ENCODED_ID or official TENCENT_DOCS_FILE_ID, "
                "then configure TENCENT_DOCS_SHEET_ID or TENCENT_DOCS_SHEET_TITLE if multiple sheets exist."
            ),
        }

    def validate_api_endpoints(self, *, write: bool) -> None:
        """检查表格读写 endpoint，未配置时不发起任何真实请求。"""

        missing_fields = self._missing_sheetbook_fields()
        if missing_fields:
            diagnosis = self.build_endpoint_diagnosis(operation="export" if write else "import")
            raise TencentDocsEndpointConfigError(
                "Tencent Docs sheetbook API requires TENCENT_DOCS_BOOK_ID; configure it or convert FILE_ID first.",
                diagnosis=diagnosis,
            )

    def _not_implemented(self) -> None:
        raise TencentDocsApiNotImplementedError("腾讯文档 API 尚未配置或尚未实现")

    def build_headers(self) -> dict[str, str]:
        """构造腾讯文档 OpenAPI 请求头，不返回或打印任何 token 明文。"""

        config = get_tencent_docs_config()
        self.validate_config()
        return {
            "Access-Token": config["access_token"],
            "Client-Id": config["client_id"],
            "Open-Id": config["open_id"],
            "Content-Type": "application/json",
        }

    def build_api_url(self, path: str) -> str:
        """根据配置集中构造腾讯文档 API URL。"""

        config = get_tencent_docs_config()
        base_url = config["api_base_url"].rstrip("/") + "/"
        return urljoin(base_url, path.lstrip("/"))

    def build_sheetbook_path(self, book_id: str, suffix: str) -> str:
        """按官方 sheetbook 路径构造 OpenAPI path。"""

        encoded_book_id = quote(book_id, safe="")
        return f"/openapi/sheetbook/v2/{encoded_book_id}/{suffix.lstrip('/')}"

    def extract_book_id_from_response(self, response_data: Any) -> str | None:
        """从 converter/metadata 响应中尽量提取 bookID。"""

        if isinstance(response_data, dict):
            for key in ("bookID", "bookId", "book_id", "bookid"):
                value = response_data.get(key)
                if value:
                    return str(value)
            for value in response_data.values():
                found = self.extract_book_id_from_response(value)
                if found:
                    return found
        if isinstance(response_data, list):
            for item in response_data:
                found = self.extract_book_id_from_response(item)
                if found:
                    return found
        return None

    def extract_file_id_from_response(self, response_data: Any) -> str | None:
        """从 converter 响应中递归提取官方 fileID。"""

        if isinstance(response_data, dict):
            for key in ("fileID", "fileId", "file_id", "fileid", "id"):
                value = response_data.get(key)
                if value:
                    return str(value)
            for value in response_data.values():
                found = self.extract_file_id_from_response(value)
                if found:
                    return found
        if isinstance(response_data, list):
            for item in response_data:
                found = self.extract_file_id_from_response(item)
                if found:
                    return found
        return None

    def resolve_file_id(self) -> str:
        """解析 OpenAPI 可用的官方 fileID。

        若配置的是 docs.qq.com/sheet/{encodedID} 中的网页 encoded id，则先调用官方 converter：
        GET /openapi/drive/v2/util/converter?type=2&value={encodedID}
        """

        config = get_tencent_docs_config()
        if self._resolved_file_id:
            return self._resolved_file_id
        if config["file_id"]:
            self._resolved_file_id = config["file_id"]
            return self._resolved_file_id
        if not config["encoded_id"]:
            diagnosis = self.build_endpoint_diagnosis(operation="resolve_file_id")
            raise TencentDocsEndpointConfigError(
                "缺少 TENCENT_DOCS_FILE_ID 或 TENCENT_DOCS_ENCODED_ID，无法解析腾讯文档官方 fileID。",
                diagnosis=diagnosis,
            )

        converter_path = "/openapi/drive/v2/util/converter"
        try:
            response_data = self.tencent_api_request(
                method="GET",
                path=converter_path,
                params={"type": 2, "value": config["encoded_id"]},
            )
        except TencentDocsApiError as exc:
            diagnosis = self.build_endpoint_diagnosis(
                operation="resolve_file_id",
                attempted_endpoint=converter_path,
                attempts=[{"path": converter_path, "error": str(exc)[:500]}],
            )
            raise TencentDocsEndpointConfigError(
                "腾讯文档 encodedID 转换 fileID 失败，请检查 access_token/open_id 权限或手动配置官方 TENCENT_DOCS_FILE_ID。",
                diagnosis=diagnosis,
            ) from exc

        file_id = self.extract_file_id_from_response(response_data)
        if not file_id:
            diagnosis = self.build_endpoint_diagnosis(
                operation="resolve_file_id",
                attempted_endpoint=converter_path,
                attempts=[{"path": converter_path, "result": "no_file_id_in_response"}],
            )
            raise TencentDocsEndpointConfigError(
                "腾讯文档 converter 响应中未找到官方 fileID，请手动配置 TENCENT_DOCS_FILE_ID。",
                diagnosis=diagnosis,
            )
        self._resolved_file_id = file_id
        return file_id

    def resolve_book_id(self) -> str:
        """获取 sheetbook API 需要的 bookID；缺失时返回可操作诊断。"""

        config = get_tencent_docs_config()
        if self._resolved_book_id:
            return self._resolved_book_id
        if config["book_id"]:
            self._resolved_book_id = config["book_id"]
            return self._resolved_book_id

        # 官方文档中 sheetbook bookID 与 drive fileID 可能一致；如果未显式配置 BOOK_ID，
        # 先使用已经解析到的官方 fileID 作为 bookID，避免把网页 encoded id 直接当成 bookID。
        self._resolved_book_id = self.resolve_file_id()
        return self._resolved_book_id

    def extract_sheet_list(self, response_data: Any) -> list[dict[str, Any]]:
        """从 sheets-info 响应中提取子表列表。"""

        if isinstance(response_data, list):
            return [item for item in response_data if isinstance(item, dict)]
        if not isinstance(response_data, dict):
            return []
        for key in (
            "properties",
            "sheets",
            "sheetList",
            "sheet_list",
            "records",
            "subSheets",
            "sheetInfo",
        ):
            value = response_data.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        data = response_data.get("data")
        if data is not response_data:
            return self.extract_sheet_list(data)
        return []

    def get_sheet_identifier(self, sheet: dict[str, Any], fallback: str) -> str:
        """优先使用 sheets-info 返回的 sheetID/tabID，取不到时回退到 sheet 名称。"""

        for key in ("sheetID", "sheetId", "sheet_id", "tabID", "tabId", "tab_id", "id"):
            value = sheet.get(key)
            if value:
                return str(value)
        return fallback

    def get_sheet_title(self, sheet: dict[str, Any]) -> str:
        """从腾讯文档 sheet 信息中提取标题。"""

        for key in ("title", "name", "sheetName", "sheet_name"):
            value = sheet.get(key)
            if value:
                return str(value)
        return ""

    def get_sheet_row_count(self, sheet: dict[str, Any]) -> int | None:
        """从多种腾讯文档响应结构中提取行数。"""

        for key in ("rowCount", "row_count", "rows", "rowNum"):
            value = sheet.get(key)
            if value is not None:
                try:
                    return int(value)
                except (TypeError, ValueError):
                    return None
        nested = sheet.get("gridProperties") or sheet.get("properties") or {}
        if isinstance(nested, dict) and nested is not sheet:
            return self.get_sheet_row_count(nested)
        return None

    def get_sheet_column_count(self, sheet: dict[str, Any]) -> int | None:
        """从多种腾讯文档响应结构中提取列数。"""

        for key in ("columnCount", "colCount", "column_count", "columns", "columnNum"):
            value = sheet.get(key)
            if value is not None:
                try:
                    return int(value)
                except (TypeError, ValueError):
                    return None
        nested = sheet.get("gridProperties") or sheet.get("properties") or {}
        if isinstance(nested, dict) and nested is not sheet:
            return self.get_sheet_column_count(nested)
        return None

    def extract_sheet_candidates(self, response_data: Any) -> list[dict[str, Any]]:
        """提取前端可展示的 sheetID 候选列表。"""

        candidates: list[dict[str, Any]] = []
        seen: set[str] = set()
        for sheet in self.extract_sheet_list(response_data):
            sheet_id = self.get_sheet_identifier(sheet, "")
            if not sheet_id or sheet_id in seen:
                continue
            seen.add(sheet_id)
            candidates.append(
                {
                    "sheetId": sheet_id,
                    "id": sheet.get("id") if isinstance(sheet, dict) else None,
                    "title": self.get_sheet_title(sheet),
                    "name": sheet.get("name") if isinstance(sheet, dict) else None,
                    "rowCount": self.get_sheet_row_count(sheet),
                    "columnCount": self.get_sheet_column_count(sheet),
                    "raw": {k: str(v) for k, v in sheet.items()} if isinstance(sheet, dict) else str(sheet),
                }
            )
        return candidates

    def list_spreadsheet_sheets(self) -> list[dict[str, Any]]:
        """调用官方 spreadsheet v3 文件信息接口获取 sheetId/title。"""

        if self._resolved_sheets is not None:
            return self._resolved_sheets
        file_id = self.resolve_file_id()
        path = f"/openapi/spreadsheet/v3/files/{quote(file_id, safe='')}"
        response_data = self.tencent_api_request(
            method="GET",
            path=path,
            params={"concise": 1},
        )
        self._resolved_sheets = self.extract_sheet_list(response_data)
        return self._resolved_sheets

    def tencent_api_raw_request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        json_body: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """调试用原始请求，不抛弃腾讯侧返回体，也不返回任何敏感请求头。"""

        url = self.build_api_url(path)
        try:
            response = requests.request(
                method=method,
                url=url,
                headers=self.build_headers(),
                params=params,
                json=json_body,
                timeout=30,
            )
        except requests.RequestException as exc:
            return {
                "success": False,
                "http_status": None,
                "path": path,
                "url": url,
                "error": str(exc),
                "raw_response": None,
            }

        try:
            response_data: Any = response.json()
        except ValueError:
            response_data = {"raw_text": response.text[:1000]}

        ret_value = None
        code_value = None
        msg_value = None
        message_value = None
        data_value = None
        if isinstance(response_data, dict):
            ret_value = response_data.get("ret")
            code_value = response_data.get("code") or response_data.get("errcode")
            msg_value = response_data.get("msg") or response_data.get("errmsg")
            message_value = response_data.get("message")
            data_value = response_data.get("data")

        business_code = code_value if code_value is not None else ret_value
        business_ok = business_code in (None, 0, "0")
        return {
            "success": response.status_code < 400 and business_ok,
            "http_status": response.status_code,
            "path": path,
            "url": url,
            "ret": ret_value,
            "code": code_value,
            "msg": msg_value,
            "message": message_value,
            "data": data_value,
            "raw_response": response_data,
        }

    def get_sheets_debug_info(self) -> dict[str, Any]:
        """返回 spreadsheet 文件信息原始响应和 sheetID 候选，供前端调试配置。"""

        config = get_tencent_docs_config()
        file_id = self.resolve_file_id()
        path = f"/openapi/spreadsheet/v3/files/{quote(file_id, safe='')}"
        raw_result = self.tencent_api_raw_request(
            method="GET",
            path=path,
            params={"concise": 1},
        )
        candidates = self.extract_sheet_candidates(raw_result.get("raw_response"))
        configured_sheet_id = config["sheet_id"]
        candidate_sheet_ids = {c["sheetId"] for c in candidates}
        sheet_id_warning = None
        if configured_sheet_id and configured_sheet_id not in candidate_sheet_ids:
            sheet_id_warning = (
                f"当前配置的 sheetID ({configured_sheet_id}) 不在腾讯 API 返回的 sheet 列表中。"
                f"可用 sheetID: {sorted(candidate_sheet_ids)}。"
                f"该值可能是网页 tab ID，不是 OpenAPI sheetID。"
            )
        return {
            "success": raw_result["success"],
            "file_id": file_id,
            "configured_sheet_id": configured_sheet_id,
            "sheet_id_warning": sheet_id_warning,
            "sheet_id_candidates": candidates,
            "raw_response": raw_result,
        }

    def resolve_sheet(self) -> dict[str, Any]:
        """解析导入/导出所需的 sheetID。

        优先使用 TENCENT_DOCS_SHEET_ID；未配置时读取 spreadsheet v3 文件信息。
        如果只有一个 sheet，自动选择；多个 sheet 时优先匹配 TENCENT_DOCS_SHEET_TITLE。
        旧 TENCENT_DOCS_TAB_ID 只用于展示兼容，不默认当成 OpenAPI sheetID。
        """

        config = get_tencent_docs_config()
        if self._resolved_sheet:
            return self._resolved_sheet
        if config["sheet_id"]:
            title = config["sheet_title"]
            if not title:
                try:
                    for sheet in self.list_spreadsheet_sheets():
                        if self.get_sheet_identifier(sheet, "") == config["sheet_id"]:
                            title = self.get_sheet_title(sheet)
                            break
                except Exception:
                    title = ""
            self._resolved_sheet = {
                "sheet_id": config["sheet_id"],
                "title": title or config["sheet_id"],
            }
            return self._resolved_sheet

        sheets = self.list_spreadsheet_sheets()
        if not sheets:
            diagnosis = self.build_endpoint_diagnosis(operation="resolve_sheet_id")
            raise TencentDocsEndpointConfigError(
                "未能从腾讯文档文件信息中读取 sheetId，请在后端配置 TENCENT_DOCS_SHEET_ID。",
                diagnosis=diagnosis,
            )
        if len(sheets) == 1:
            sheet = sheets[0]
            self._resolved_sheet = {
                "sheet_id": self.get_sheet_identifier(sheet, ""),
                "title": self.get_sheet_title(sheet) or self.get_sheet_identifier(sheet, ""),
            }
            return self._resolved_sheet

        if config["sheet_title"]:
            for sheet in sheets:
                if self.get_sheet_title(sheet) == config["sheet_title"]:
                    self._resolved_sheet = {
                        "sheet_id": self.get_sheet_identifier(sheet, ""),
                        "title": self.get_sheet_title(sheet),
                    }
                    return self._resolved_sheet

        diagnosis = self.build_endpoint_diagnosis(
            operation="resolve_sheet_id",
            attempts=[
                {
                    "result": "multiple_sheets_need_manual_sheet_id",
                    "sheet_titles": [self.get_sheet_title(sheet) for sheet in sheets],
                }
            ],
        )
        raise TencentDocsEndpointConfigError(
            "腾讯文档包含多个 sheet，无法自动确定目标 sheet，请配置 TENCENT_DOCS_SHEET_ID 或 TENCENT_DOCS_SHEET_TITLE。",
            diagnosis=diagnosis,
        )

    def get_target_sheets_for_year(self, year: int) -> list[dict[str, str]]:
        """获取某一年需要导入/导出的目标 sheet 列表。"""

        config = get_tencent_docs_config()
        sheets = self.list_spreadsheet_sheets()
        if config["sheet_id"]:
            resolved = self.resolve_sheet()
            return [{"sheet_id": resolved["sheet_id"], "title": resolved.get("title") or str(year)}]

        matched: list[dict[str, str]] = []
        for sheet in sheets:
            title = self.get_sheet_title(sheet)
            period = parse_sheet_year_month(title) if title else None
            if period and period[0] == year:
                matched.append({"sheet_id": self.get_sheet_identifier(sheet, ""), "title": title})
        if matched:
            return matched

        resolved = self.resolve_sheet()
        return [{"sheet_id": resolved["sheet_id"], "title": resolved.get("title") or str(year)}]

    def resolve_sheet_id_for_month(
        self,
        year: int,
        month: int,
        *,
        force_env_sheet: bool = False,
    ) -> dict[str, Any]:
        """Resolve sheetID for a specific year/month by matching sheet titles.

        Resolution order:
        1. TENCENT_DOCS_SHEET_MAP_JSON if configured and key present
        2. Live spreadsheet sheet list, matching by title pattern
        3. Only if force_env_sheet=True: TENCENT_DOCS_SHEET_ID as last resort

        Returns dict with sheet_id, sheet_title, source, candidates.
        Raises TencentDocsEndpointConfigError if no match found.
        """

        import json as _json

        config = get_tencent_docs_config()
        candidates = self.extract_sheet_candidates(self.list_spreadsheet_sheets())
        key = f"{year}.{month}"

        # 1. Sheet map from env
        sheet_map_raw = (settings.TENCENT_DOCS_SHEET_MAP_JSON or "").strip()
        if sheet_map_raw:
            try:
                sheet_map = _json.loads(sheet_map_raw)
            except _json.JSONDecodeError:
                sheet_map = {}
            if isinstance(sheet_map, dict) and key in sheet_map:
                sheet_id = str(sheet_map[key])
                return {
                    "sheet_id": sheet_id,
                    "sheet_title": key,
                    "source": "sheet_map",
                    "candidates": candidates,
                }

        # 2. Match by title from live sheet list
        title_aliases = [
            key,                              # 2026.5
            f"{year}.{month:02d}",            # 2026.05
            f"{month}月",                     # 5月
            f"{year}年{month}月",             # 2026年5月
        ]
        for sheet in self.list_spreadsheet_sheets():
            title = self.get_sheet_title(sheet)
            for alias in title_aliases:
                if title == alias:
                    return {
                        "sheet_id": self.get_sheet_identifier(sheet, ""),
                        "sheet_title": title,
                        "source": "sheet_list",
                        "candidates": candidates,
                    }
        # Also try parse_sheet_year_month for any sheet
        for sheet in self.list_spreadsheet_sheets():
            title = self.get_sheet_title(sheet)
            period = parse_sheet_year_month(title) if title else None
            if period and period[0] == year and period[1] == month:
                return {
                    "sheet_id": self.get_sheet_identifier(sheet, ""),
                    "sheet_title": title,
                    "source": "sheet_list_parsed",
                    "candidates": candidates,
                }

        # 3. Debug/legacy: force_env_sheet fallback
        env_sheet_id = config["sheet_id"]
        if force_env_sheet and env_sheet_id:
            return {
                "sheet_id": env_sheet_id,
                "sheet_title": config.get("sheet_title") or key,
                "source": "env_sheet_id",
                "candidates": candidates,
            }

        raise TencentDocsEndpointConfigError(
            f"未找到 {year} 年 {month} 月对应的腾讯文档 sheet",
            diagnosis={
                "operation": "resolve_sheet_id_for_month",
                "year": year,
                "month": month,
                "search_key": key,
                "tried_aliases": title_aliases,
                "sheet_candidates": candidates,
                "tip": "请确认腾讯文档中 sheet 标题格式为 {year}.{month}（例如 2026.5），或在 TENCENT_DOCS_SHEET_MAP_JSON 中显式映射。",
            },
        )

    def tencent_api_request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        json_body: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """集中调用腾讯文档 OpenAPI。

        错误信息包含 URL、HTTP 状态码和腾讯侧返回摘要，但不会包含 access_token。
        """

        url = self.build_api_url(path)
        try:
            response = requests.request(
                method=method,
                url=url,
                headers=self.build_headers(),
                params=params,
                json=json_body,
                timeout=30,
            )
        except requests.RequestException as exc:
            raise TencentDocsApiError(f"腾讯文档 API 请求失败：url={url}，error={exc}") from exc

        try:
            response_data = response.json()
        except ValueError:
            response_data = {"raw_text": response.text[:500]}

        if response.status_code >= 400:
            raise TencentDocsApiError(
                f"腾讯文档 API 返回错误：url={url}，status_code={response.status_code}，"
                f"response={response_data}",
                detail={
                    "request_url": url,
                    "http_status": response.status_code,
                    "code": response_data.get("code") or response_data.get("ret") or response_data.get("errcode"),
                    "msg": response_data.get("msg") or response_data.get("errmsg"),
                    "raw_response": str(response_data)[:2000],
                },
            )

        if isinstance(response_data, dict):
            code = response_data.get("code") or response_data.get("ret") or response_data.get("errcode")
            if code not in (None, 0, "0"):
                raise TencentDocsApiError(
                    f"腾讯文档 API 业务错误：url={url}，code={code}，response={response_data}",
                    detail={
                        "request_url": url,
                        "http_status": response.status_code,
                        "code": code,
                        "msg": response_data.get("msg") or response_data.get("errmsg"),
                        "raw_response": str(response_data)[:2000],
                    },
                )
            return response_data

        return {"data": response_data}

    def get_status(self) -> dict[str, Any]:
        config = get_tencent_docs_config()
        token = self._latest_token()
        token_expiry_info = get_token_expiry_info(self.db) if self.db else {
            "token_expires_at": None,
            "source": "none",
            "status": "unknown",
            "remaining_seconds": None,
            "remaining_text": "-",
            "warning_threshold_days": 7,
            "expiring_soon": False,
        }
        direct_token_configured = bool(config["access_token"])
        token_saved = bool(direct_token_configured or (token and token.access_token))
        oauth_token_expires_at = token.expires_at if token else None
        # Direct Token 来自腾讯文档开放生态后台，无法从本地判断过期时间；只要配置了 token 和 open_id，
        # 就认为当前具备调用条件。OAuth token 仍按数据库 expires_at 判断有效性。
        token_valid = bool(
            (direct_token_configured and config["open_id"])
            or (token_saved and oauth_token_expires_at and oauth_token_expires_at > now_beijing())
        )
        open_id_saved = bool(config["open_id"] or (token and token.open_id))
        ready_for_oauth = all(
            [config["client_id"], config["client_secret"], config["redirect_uri"]]
        )
        file_id_configured = bool(config["file_id"])
        encoded_id_configured = bool(config["encoded_id"])
        book_id_configured = bool(config["book_id"])
        sheet_id_configured = bool(config["sheet_id"])
        sheet_title_configured = bool(config["sheet_title"])
        tab_id_configured = bool(config["tab_id"])
        ready_for_direct_token = all(
            [
                config["client_id"],
                config["access_token"],
                config["open_id"],
                file_id_configured or encoded_id_configured or book_id_configured,
            ]
        )
        file_id_resolved = False
        book_id_resolved = False
        sheet_id_resolved = False
        resolved_file_id: str | None = None
        resolved_book_id: str | None = None
        resolved_sheet_id: str | None = None
        resolved_sheet_title: str | None = None
        sheet_id_candidates: list[dict[str, Any]] = []
        last_probe_error: str | None = None

        if ready_for_direct_token:
            try:
                resolved_file_id = self.resolve_file_id()
                file_id_resolved = bool(resolved_file_id)
            except Exception as exc:  # noqa: BLE001 - 状态接口需要返回可读诊断
                last_probe_error = str(exc)
            try:
                resolved_book_id = self.resolve_book_id()
                book_id_resolved = bool(resolved_book_id)
            except Exception as exc:  # noqa: BLE001
                last_probe_error = last_probe_error or str(exc)
            try:
                sheet_id_candidates = self.extract_sheet_candidates(self.list_spreadsheet_sheets())
                target_sheets = self.get_target_sheets_for_year(int(config["default_year"]))
                if target_sheets:
                    resolved_sheet_id = target_sheets[0].get("sheet_id")
                    resolved_sheet_title = target_sheets[0].get("title")
                    sheet_id_resolved = bool(resolved_sheet_id)
            except Exception as exc:  # noqa: BLE001
                last_probe_error = last_probe_error or str(exc)

        sheet_range_configured = bool(config["sheet_range"])
        read_endpoint_enabled = True
        write_endpoint_enabled = True
        ready_for_api_endpoint = bool(read_endpoint_enabled and write_endpoint_enabled)
        credentials_ready = bool(ready_for_direct_token or (ready_for_oauth and token_saved and token_valid))
        ready_for_import = bool(
            credentials_ready and file_id_resolved and sheet_id_resolved and sheet_range_configured and read_endpoint_enabled
        )
        ready_for_export = bool(
            credentials_ready and book_id_resolved and sheet_id_resolved and sheet_range_configured and write_endpoint_enabled
        )

        return {
            "mode": config["mode"],
            "client_id_configured": bool(config["client_id"]),
            "client_secret_configured": bool(config["client_secret"]),
            "redirect_uri_configured": bool(config["redirect_uri"]),
            "access_token_configured": direct_token_configured,
            "doc_id_configured": bool(config["doc_id"] or config["book_id"]),
            "doc_id": config["doc_id"] or config["book_id"] or None,
            "encoded_id_configured": encoded_id_configured,
            "encoded_id": config["encoded_id"] or None,
            "file_id_configured": file_id_configured,
            "file_id_resolved": file_id_resolved,
            "file_id": resolved_file_id or config["file_id"] or None,
            "book_id_configured": book_id_configured,
            "book_id_resolved": book_id_resolved,
            "book_id": resolved_book_id or config["book_id"] or None,
            "sheet_id_configured": sheet_id_configured,
            "sheet_id_resolved": sheet_id_resolved,
            "sheet_id": resolved_sheet_id or config["sheet_id"] or None,
            "sheet_id_candidates": sheet_id_candidates,
            "sheet_title_configured": sheet_title_configured,
            "sheet_title": resolved_sheet_title or config["sheet_title"] or None,
            "tab_id_configured": tab_id_configured,
            "tab_id": config["tab_id"] or None,
            "sheet_range_configured": sheet_range_configured,
            "sheet_range": config["sheet_range"],
            "read_range": config["read_range"],
            "write_range": config["write_range"],
            "template_type": config["template_type"],
            "active_month": config["active_month"],
            "default_year": config["default_year"],
            "auth_mode": "direct_token" if direct_token_configured else "oauth",
            "ready_for_direct_token": ready_for_direct_token,
            "read_endpoint_enabled": read_endpoint_enabled,
            "write_endpoint_enabled": write_endpoint_enabled,
            "sheet_read_endpoint_configured": read_endpoint_enabled,
            "sheet_write_endpoint_configured": write_endpoint_enabled,
            "ready_for_api_endpoint": ready_for_api_endpoint,
            "ready_for_import": ready_for_import,
            "ready_for_export": ready_for_export,
            "last_probe_error": last_probe_error,
            "token_saved": token_saved,
            "token_valid": token_valid,
            "token_expires_at": token_expiry_info.get("token_expires_at")
            or (oauth_token_expires_at.isoformat() if oauth_token_expires_at else None),
            "token_expiry_source": token_expiry_info.get("source"),
            "token_expiry_status": token_expiry_info.get("status"),
            "token_remaining_seconds": token_expiry_info.get("remaining_seconds"),
            "token_remaining_text": token_expiry_info.get("remaining_text"),
            "token_expiring_soon": bool(token_expiry_info.get("expiring_soon")),
            "token_expiry_warning_threshold_days": token_expiry_info.get("warning_threshold_days"),
            "open_id_saved": open_id_saved,
            "ready_for_oauth": ready_for_oauth,
            "ready_for_sync": bool(ready_for_import or ready_for_export),
        }

    def validate_config(self) -> None:
        config = get_tencent_docs_config()
        if config["access_token"]:
            missing_fields = self._missing_direct_token_fields()
        else:
            missing_fields = self._missing_oauth_fields()
        if not (config["file_id"] or config["encoded_id"] or config["book_id"]):
            missing_fields.append("TENCENT_DOCS_FILE_ID 或 TENCENT_DOCS_ENCODED_ID")
        if missing_fields:
            raise TencentDocsConfigError(
                "缺少腾讯文档真实同步配置：" + "、".join(missing_fields)
            )

    def build_oauth_url(self, state: str | None = None) -> str:
        """生成 OAuth 授权 URL；配置不足时抛出清晰错误。"""

        missing_fields = self._missing_oauth_fields()
        if missing_fields:
            raise TencentDocsConfigError(
                "缺少腾讯文档 OAuth 配置：" + "、".join(missing_fields)
            )

        config = get_tencent_docs_config()
        params = {
            "client_id": config["client_id"],
            "redirect_uri": config["redirect_uri"],
            "response_type": "code",
            "scope": settings.TENCENT_DOC_SCOPE,
        }
        if state:
            params["state"] = state
        return f"{settings.TENCENT_DOC_OAUTH_AUTHORIZE_URL}?{urlencode(params)}"

    def handle_oauth_callback(self, code: str, state: str | None = None) -> None:
        _ = code, state
        self._not_implemented()

    def refresh_token_if_needed(self) -> None:
        self._not_implemented()

    def list_sheets(self) -> list[dict[str, Any]]:
        return self.list_spreadsheet_sheets()

    def read_sheet_range(self, sheet_id: str, range_name: str) -> list[list[Any]]:
        file_id = self.resolve_file_id()
        actual_sheet_id = sheet_id or self.resolve_sheet()["sheet_id"]
        actual_range = range_name or get_tencent_docs_config()["read_range"]
        encoded_file_id = quote(file_id, safe="")
        encoded_sheet_id = quote(actual_sheet_id, safe="")
        encoded_range = quote(actual_range, safe="")
        response_data = self.tencent_api_request(
            method="GET",
            path=f"/openapi/spreadsheet/v3/files/{encoded_file_id}/{encoded_sheet_id}/{encoded_range}",
        )
        return self.extract_values_from_response(response_data)

    def write_sheet_range(self, sheet_id: str, range_name: str, values: list[list[Any]]) -> dict[str, Any]:
        """Write values to a sheet range via sheetbook v2 API.

        Returns the raw result dict so callers can inspect the full response.
        """

        book_id = self.resolve_book_id()
        actual_sheet_id = sheet_id or self.resolve_sheet()["sheet_id"]
        actual_range = range_name or get_tencent_docs_config()["write_range"]
        encoded_range = quote(f"{actual_sheet_id}!{actual_range}", safe="!")
        api_path = self.build_sheetbook_path(book_id, f"values/{encoded_range}")
        payload = {"values": values}

        try:
            result = self.tencent_api_request(
                method="PUT",
                path=api_path,
                json_body=payload,
            )
            return {"success": True, "data": result}
        except TencentDocsApiError as exc:
            exc.detail["book_id"] = book_id
            exc.detail["sheet_id"] = actual_sheet_id
            exc.detail["range"] = actual_range
            exc.detail["api_path"] = api_path
            exc.detail["request_url"] = self.build_api_url(api_path)
            raise

    def extract_values_from_response(self, response_data: Any) -> list[list[Any]]:
        """尽量从腾讯文档 API 响应中提取二维表格 values。"""

        if isinstance(response_data, list):
            return response_data
        if not isinstance(response_data, dict):
            return []
        for key in ("values", "rows", "data", "records"):
            value = response_data.get(key)
            if isinstance(value, list):
                if value and all(isinstance(row, dict) for row in value):
                    return [list(row.values()) for row in value]
                return value
            if isinstance(value, dict):
                nested_values = self.extract_values_from_response(value)
                if nested_values:
                    return nested_values
        return []

    def convert_tencent_rows_to_normalized_records(
        self,
        sheet_name: str,
        rows: list[list[Any]],
    ) -> tuple[list[NormalizedInventoryRecord], SyncImportResult]:
        """把腾讯文档二维行列数据转换为标准库存流水记录。

        解析规则与 Excel 导入保持一致：第 2 行试剂名、第 3 行子列、第 4 行开始日流水。
        """

        result = SyncImportResult()
        period = parse_sheet_year_month(sheet_name)
        if period is None:
            result.add_error(sheet_name, None, None, f"无法从 sheet 名识别年月：{sheet_name}")
            return [], result

        year, month = period
        normalized_records: list[NormalizedInventoryRecord] = []
        if len(rows) < 3:
            return normalized_records, result

        max_columns = max((len(row) for row in rows), default=0)
        padded_rows = [list(row) + [None] * (max_columns - len(row)) for row in rows]
        header_row = padded_rows[1]
        groups: list[tuple[str, int]] = []
        for column_index, value in enumerate(header_row):
            if column_index == 0:
                continue
            reagent_name = clean_text(value)
            if reagent_name and reagent_name not in {"操作", "数量", "操作人", "日期"}:
                groups.append((reagent_name, column_index))

        for row_index in range(3, min(len(padded_rows), 34)):
            excel_row = row_index + 1
            row = padded_rows[row_index]
            try:
                operation_time = parse_excel_row_date(sheet_name, excel_row, row[0])
            except ValueError as exc:
                if self.tencent_row_has_operation(row, groups):
                    result.add_error(sheet_name, excel_row, None, str(exc))
                continue

            for reagent_name, column_index in groups:
                if column_index + 2 >= len(row):
                    continue

                operation_text = clean_text(row[column_index])
                quantity_cell = row[column_index + 1]
                raw_operator_text = clean_text(row[column_index + 2])
                if not operation_text and is_blank(quantity_cell) and not raw_operator_text:
                    continue
                operator_text = raw_operator_text or "-"
                if operation_text not in OPERATION_MAPPING:
                    result.add_error(sheet_name, excel_row, reagent_name, "操作必须为入库、领取或出库")
                    continue
                try:
                    quantity = clean_quantity(quantity_cell)
                except ValueError as exc:
                    result.add_error(sheet_name, excel_row, reagent_name, str(exc))
                    continue

                normalized_records.append(
                    NormalizedInventoryRecord(
                        year=year,
                        month=month,
                        event_date=operation_time.date(),
                        reagent_name=reagent_name,
                        operation_text=operation_text,
                        operation_type=OPERATION_MAPPING[operation_text],
                        quantity=quantity,
                        operator=operator_text,
                        remark=f"腾讯文档导入：{sheet_name}!R{excel_row}C{column_index + 1}",
                        source=self.source,
                        source_sheet=sheet_name,
                        source_row=excel_row,
                        source_col=column_index + 1,
                        operation_time=operation_time,
                    )
                )
        return normalized_records, result

    def tencent_row_has_operation(self, row: list[Any], groups: list[tuple[str, int]]) -> bool:
        """判断腾讯文档某行是否包含任意操作记录。"""

        for _, column_index in groups:
            if column_index + 2 >= len(row):
                continue
            if clean_text(row[column_index]) or not is_blank(row[column_index + 1]) or clean_text(row[column_index + 2]):
                return True
        return False

    def fetch_tencent_doc_rows(self, year: int) -> list[NormalizedInventoryRecord]:
        """从腾讯文档读取 12 个历史模板 sheet，并转换为标准库存流水。"""

        all_records: list[NormalizedInventoryRecord] = []
        self._last_parse_result = SyncImportResult()
        read_success = 0
        last_error: str | None = None
        for sheet in self.get_target_sheets_for_year(year):
            sheet_id = sheet["sheet_id"]
            sheet_name = sheet["title"]
            try:
                values = self.read_sheet_range(sheet_id, get_tencent_docs_config()["sheet_range"])
            except TencentDocsApiError as exc:
                last_error = str(exc)
                continue
            read_success += 1
            records, parse_result = self.convert_tencent_rows_to_normalized_records(sheet_name, values)
            all_records.extend(records)
            self._last_parse_result.failed += parse_result.failed
            self._last_parse_result.errors.extend(parse_result.errors)

        if read_success == 0 and last_error:
            raise TencentDocsApiError(
                "腾讯文档读取失败：已构造 API URL 且 token/file_id 配置存在，但 12 个 sheet 均未读取成功；"
                f"最后一次错误：{last_error}"
            )
        return all_records

    def export_inventory_to_tencent_rows(self, db: Session, year: int) -> dict[str, list[list[Any]]]:
        """复用 Excel 宽表导出逻辑，生成可写回腾讯文档的二维 values。"""

        from openpyxl import load_workbook

        file_path = export_inventory_excel(db, year=year)
        workbook = load_workbook(file_path, data_only=True)
        sheet_values: dict[str, list[list[Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows: list[list[Any]] = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(35, worksheet.max_row),
                max_col=worksheet.max_column,
                values_only=True,
            ):
                rows.append(["" if cell is None else cell for cell in row])
            sheet_values[sheet_name] = rows
        workbook.close()
        return sheet_values

    def import_records(self, db: Session, operator_id: int | None = None, **kwargs: Any) -> SyncImportResult:
        """从腾讯文档读取库存流水并复用 ImportService 写入数据库。"""

        year = int(kwargs.get("year") or get_tencent_docs_config()["default_year"])
        records = self.fetch_tencent_doc_rows(year=year)
        result = ImportService(db=db, operator_id=operator_id).import_records(records)
        parse_result = getattr(self, "_last_parse_result", SyncImportResult())
        result.failed += parse_result.failed
        result.errors = parse_result.errors + result.errors
        return result

    def export_records(self, db: Session, year: int = 2026, **kwargs: Any) -> Any:
        """把本地库存流水导出为历史模板宽表，并写回腾讯文档。"""

        _ = kwargs
        self.validate_config()
        self.resolve_book_id()
        sheet_lookup = {
            sheet["title"]: sheet["sheet_id"]
            for sheet in self.get_target_sheets_for_year(year)
            if sheet.get("title") and sheet.get("sheet_id")
        }
        sheet_values = self.export_inventory_to_tencent_rows(db=db, year=year)
        written_sheets = 0
        skipped_sheets = 0
        total_rows = 0
        for sheet_name, values in sheet_values.items():
            if not values:
                continue
            sheet_identifier = sheet_lookup.get(sheet_name)
            if not sheet_identifier:
                skipped_sheets += 1
                continue
            raise TencentDocsApiNotImplementedError(
                "旧版整块写回逻辑已禁用；正式同步必须使用增量 1x3 patch 写入。"
            )
            written_sheets += 1
            total_rows += len(values)
        return {
            "year": year,
            "written_sheets": written_sheets,
            "skipped_sheets": skipped_sheets,
            "total_rows": total_rows,
        }


    def _active_month(self, month: int | None = None) -> int:
        config = get_tencent_docs_config()
        value = int(month or config.get("active_month") or 1)
        if not 1 <= value <= 12:
            raise TencentDocsConfigError("TENCENT_DOCS_ACTIVE_MONTH must be between 1 and 12")
        return value

    def preview_import_matrix(self, year: int, month: int | None = None) -> dict[str, Any]:
        """Read active Tencent Docs range and parse it as the reagent matrix template.

        Uses raw request to capture full debug info even when the API returns errors.
        """

        config = get_tencent_docs_config()
        target_month = self._active_month(month)
        sheet = self.resolve_sheet()
        file_id = self.resolve_file_id()
        actual_sheet_id = sheet["sheet_id"]
        actual_range = config["read_range"]
        encoded_file_id = quote(file_id, safe="")
        encoded_sheet_id = quote(actual_sheet_id, safe="")
        encoded_range = quote(actual_range, safe="")
        api_path = f"/openapi/spreadsheet/v3/files/{encoded_file_id}/{encoded_sheet_id}/{encoded_range}"

        raw_result = self.tencent_api_raw_request(
            method="GET",
            path=api_path,
        )
        request_url = raw_result["url"]
        http_status = raw_result["http_status"]
        raw_response = raw_result["raw_response"]

        # Attempt to extract values from the raw response
        values: list[list[Any]] = []
        api_error = None
        if raw_result["success"] and raw_response is not None:
            values = self.extract_values_from_response(raw_response)

        if not raw_result["success"]:
            api_error = {
                "request_url": request_url,
                "http_status": http_status,
                "code": raw_result.get("code"),
                "msg": raw_result.get("msg"),
                "message": raw_result.get("message"),
            }

        if values:
            parse_result = parse_reagent_matrix(
                values,
                year=year,
                month=target_month,
                sheet_name=sheet.get("title") or f"{year}.{target_month}",
            )
            response = parse_result.to_debug_response()
        else:
            # No values parsed — return the raw response so we can debug
            response = {
                "detected_template_type": MATRIX_TEMPLATE_TYPE,
                "year": year,
                "month": target_month,
                "raw_rows_count": 0,
                "reagent_count": 0,
                "reagent_names": [],
                "parsed_records_preview": [],
                "parsed_records_count": 0,
                "invalid_records": [],
                "invalid_records_preview": [],
                "raw_values_preview": [],
                "parsed_values_shape": None,
                "matrix_row_1_preview": [],
                "matrix_row_2_preview": [],
                "matrix_row_3_preview": [],
                "detected_reagent_columns": [],
            }

        response["sheet_id"] = sheet["sheet_id"]
        response["sheet_title"] = sheet.get("title")
        response["range"] = actual_range
        response["request_url"] = request_url
        response["http_status"] = http_status
        response["raw_response"] = raw_response
        if api_error:
            response["api_error"] = api_error
        return response

    def fetch_tencent_doc_rows(
        self,
        year: int,
        month: int | None = None,
    ) -> list[NormalizedInventoryRecord]:
        """Read Tencent Docs matrix data and convert it to normalized records.

        Resolves the correct sheetID dynamically from sheet titles matching year/month.
        """

        config = get_tencent_docs_config()
        target_month = self._active_month(month)
        resolved = self.resolve_sheet_id_for_month(year, target_month)
        sheet_id = resolved["sheet_id"]
        sheet_title = resolved.get("sheet_title") or f"{year}.{target_month}"
        values = self.read_sheet_range(sheet_id, config["read_range"])
        parse_result = parse_reagent_matrix(
            values,
            year=year,
            month=target_month,
            sheet_name=sheet_title,
        )
        self._last_matrix_parse = parse_result
        self._last_resolved_sheet_id = sheet_id
        self._last_resolved_sheet_title = sheet_title
        parse_errors = SyncImportResult()
        parse_errors.failed = len(parse_result.invalid_records)
        parse_errors.errors = parse_result.invalid_records
        self._last_parse_result = parse_errors
        return parse_result.records

    def preview_export_matrix(self, db: Session, year: int, month: int | None = None) -> dict[str, Any]:
        """Build patch preview for non-destructive Tencent Docs append sync.

        Resolves the correct sheetID dynamically from sheet titles.
        """

        config = get_tencent_docs_config()
        target_month = self._active_month(month)
        resolved = self.resolve_sheet_id_for_month(year, target_month)
        sheet_id = resolved["sheet_id"]
        sheet_title = resolved.get("sheet_title") or f"{year}.{target_month}"

        preview = build_reagent_matrix_patches_from_db(
            db=db,
            year=year,
            month=target_month,
            sheet_id=sheet_id,
        )
        preview["read_range"] = config["read_range"]
        preview["data_area_range"] = config["write_range"]
        preview["write_range"] = "incremental_patches"
        preview["sheet_id"] = sheet_id
        preview["sheet_title"] = sheet_title

        merged_patches: list[dict[str, Any]] = []
        failed_patches: list[dict[str, Any]] = []
        existing_cells_count = 0
        appended_tuples_count = 0
        skipped_duplicate_count = 0
        changed_cells_count = 0

        for patch in preview.get("patches", []):
            try:
                existing_values = self.read_sheet_range(sheet_id, patch["range"])
            except TencentDocsApiError as exc:
                failed_patches.append(
                    {
                        "day": patch.get("day"),
                        "reagent_name": patch.get("reagent_name"),
                        "range": patch.get("full_range") or f"{sheet_id}!{patch.get('range')}",
                        "reason": "读取腾讯文档已有三联单元格失败，已跳过该 patch，避免覆盖原表。",
                        "error": exc.detail or {"error": str(exc)},
                    }
                )
                continue
            merged_patch = merge_patch_with_existing(patch, existing_values)
            before = merged_patch.get("before", {})
            existing_cells_count += sum(
                1 for key in ("operation", "quantity", "operator") if str(before.get(key) or "").strip()
            )
            appended_tuples_count += int(merged_patch.get("appended_tuples_count") or 0)
            skipped_duplicate_count += int(merged_patch.get("skipped_duplicate_count") or 0)
            changed_cells_count += int(merged_patch.get("changed_cells_count") or 0)
            merged_patches.append(merged_patch)

        will_write_patches = [patch for patch in merged_patches if patch.get("will_write")]
        preview["patches"] = merged_patches
        preview["failed_patches"] = failed_patches
        preview["failed_read_count"] = len(failed_patches)
        preview["existing_cells_count"] = existing_cells_count
        preview["appended_tuples_count"] = appended_tuples_count
        preview["skipped_duplicate_count"] = skipped_duplicate_count
        preview["skipped_duplicate_tuples_count"] = skipped_duplicate_count
        preview["changed_cells_count"] = changed_cells_count
        preview["will_write_patch_count"] = len(will_write_patches)
        preview["failed_patch_count"] = len(failed_patches)
        preview["preview_changes"] = will_write_patches[:10]
        preview["will_call_tencent_api"] = bool(will_write_patches)
        preview["values_preview"] = []
        return preview

    def debug_write_cell(self, year: int = 2026, month: int | None = None) -> dict[str, Any]:
        """Safe write-cell connectivity test.

        Uses the configured test sheetID and range (default 000001!A1:A1).
        Does NOT require TENCENT_DOCS_SHEET_ID / TENCENT_DOCS_SHEET_TITLE.
        Writes exactly one cell — never touches A1:BF37 or any data region.
        """

        from datetime import datetime as dt

        config = get_tencent_docs_config()
        # Resolve only credentials and bookID — NOT sheet resolution
        file_id = self.resolve_file_id()
        book_id = self.resolve_book_id()

        # Test target: fixed config or hardcoded default, independent of formal sync
        raw_sheet_id = (config["write_cell_test_sheet_id"] or "000001").strip()
        raw_range = (config["write_cell_test_range"] or "A1:A1").strip()
        # If user accidentally pasted a full range like "000001!A1:A1", extract just the cell part
        if "!" in raw_range:
            raw_sheet_id, raw_range = raw_range.split("!", 1)
            raw_sheet_id = raw_sheet_id.strip()
            raw_range = raw_range.strip()
        sheet_id = raw_sheet_id or "000001"
        test_range = raw_range or "A1:A1"

        timestamp = dt.now().strftime("%Y-%m-%d %H:%M:%S")
        test_value = f"sync_test_{timestamp}"

        primary_attempt = self._try_write_cell(
            book_id=book_id,
            sheet_id=sheet_id,
            label=f"Write-Cell test: {sheet_id}!{test_range}",
            test_value=test_value,
            range_name=test_range,
        )
        success = bool(primary_attempt.get("success"))
        return {
            "success": success,
            "book_id_used": book_id,
            "file_id_used": file_id,
            "sheet_id_used": sheet_id,
            "range_used": f"{sheet_id}!{test_range}",
            "test_value": test_value,
            "request_url": primary_attempt.get("request_url"),
            "http_status": primary_attempt.get("http_status"),
            "raw_response": primary_attempt.get("raw_response"),
            "any_success": success,
            "primary_attempt": primary_attempt,
            "attempts": [primary_attempt],
            "message": (
                "Write-Cell 测试成功，请打开腾讯文档第一个 sheet，查看左上角 A1 单元格是否出现测试值。"
                if success
                else (
                    "Write-Cell 测试目标 "
                    + f"sheetID={sheet_id} 可能不存在，或当前文档不支持该 sheetID。"
                    + "请通过探测 Sheet 列表确认第一个 sheet 的 sheetID，"
                    + "或修改 TENCENT_DOCS_WRITE_CELL_TEST_SHEET_ID。"
                )
            ),
        }

    def _try_write_cell(
        self,
        book_id: str,
        sheet_id: str,
        label: str,
        test_value: str,
        range_name: str = "A1:A1",
    ) -> dict[str, Any]:
        """Attempt a single-cell write to the specified range."""

        encoded_range = quote(f"{sheet_id}!{range_name}", safe="!")
        api_path = self.build_sheetbook_path(book_id, f"values/{encoded_range}")
        request_url = self.build_api_url(api_path)
        payload = {"values": [[test_value]]}

        raw_result = self.tencent_api_raw_request(
            method="PUT",
            path=api_path,
            json_body=payload,
        )

        return {
            "label": label,
            "book_id": book_id,
            "sheet_id": sheet_id,
            "range": f"{sheet_id}!{range_name}",
            "request_url": request_url,
            "request_body_preview": {"values": [["sync_test_<timestamp>"]]},
            "http_status": raw_result.get("http_status"),
            "success": raw_result.get("success", False),
            "code": raw_result.get("code"),
            "msg": raw_result.get("msg"),
            "raw_response": raw_result.get("raw_response"),
        }

    def export_inventory_to_tencent_rows(
        self,
        db: Session,
        year: int,
        month: int | None = None,
    ) -> dict[str, Any]:
        """Build Tencent Docs matrix values from the database."""

        return self.preview_export_matrix(db=db, year=year, month=month)

    def import_all_months(self, db: Session, operator_id: int, year: int) -> dict[str, Any]:
        """Import from all 12 months of a year, resolving sheetID dynamically per month."""

        per_month: list[dict[str, Any]] = []
        total_inserted = 0
        total_skipped = 0
        total_failed = 0
        overall_status = "success"

        for month in range(1, 13):
            month_result: dict[str, Any] = {
                "month": month,
                "sheet_id": "",
                "sheet_title": "",
                "status": "success",
                "inserted_count": 0,
                "skipped_count": 0,
                "failed_count": 0,
                "message": "",
            }
            try:
                resolved = self.resolve_sheet_id_for_month(year, month)
                month_result["sheet_id"] = resolved["sheet_id"]
                month_result["sheet_title"] = resolved["sheet_title"]
            except TencentDocsEndpointConfigError as exc:
                month_result["status"] = "sheet_not_found"
                month_result["message"] = str(exc)
                per_month.append(month_result)
                overall_status = "partial_success"
                continue

            try:
                result = self.import_records(db=db, operator_id=operator_id, year=year, month=month)
                month_result["inserted_count"] = result.created
                month_result["skipped_count"] = result.skipped
                month_result["failed_count"] = result.failed
                month_result["message"] = result.message
                if result.failed > 0:
                    month_result["status"] = "partial_success"
                    overall_status = "partial_success"
                total_inserted += result.created
                total_skipped += result.skipped
                total_failed += result.failed
            except Exception as exc:
                month_result["status"] = "failed"
                month_result["message"] = str(exc)[:500]
                overall_status = "partial_success"

            per_month.append(month_result)

        return {
            "status": overall_status,
            "year": year,
            "mode": "all_months",
            "total_inserted": total_inserted,
            "total_skipped": total_skipped,
            "total_failed": total_failed,
            "per_month_results": per_month,
        }

    def export_all_months(self, db: Session, year: int, **kwargs: Any) -> dict[str, Any]:
        """Export to all 12 months of a year with incremental patch writes."""

        per_month: list[dict[str, Any]] = []
        total_written = 0
        total_skipped = 0
        total_failed = 0
        overall_status = "success"

        for month in range(1, 13):
            month_result: dict[str, Any] = {
                "month": month,
                "sheet_id": "",
                "sheet_title": "",
                "status": "success",
                "db_records_count": 0,
                "patch_count": 0,
                "written_patch_count": 0,
                "skipped_duplicate_count": 0,
                "failed_patch_count": 0,
                "message": "",
            }
            try:
                resolved = self.resolve_sheet_id_for_month(year, month)
                month_result["sheet_id"] = resolved["sheet_id"]
                month_result["sheet_title"] = resolved["sheet_title"]
            except TencentDocsEndpointConfigError as exc:
                month_result["status"] = "sheet_not_found"
                month_result["message"] = str(exc)
                per_month.append(month_result)
                overall_status = "partial_success"
                continue

            try:
                export_kwargs = {**kwargs, "month": month}
                result = self._export_matrix_current_month(db=db, year=year, **export_kwargs)
                month_result["db_records_count"] = result.get("db_records_count", 0)
                month_result["patch_count"] = result.get("patch_count", 0)
                month_result["written_patch_count"] = result.get("written_patch_count", 0)
                month_result["skipped_duplicate_count"] = result.get("skipped_duplicate_count", 0)
                month_result["failed_patch_count"] = result.get("failed_patch_count", 0)
                month_result["message"] = result.get("message", "")
                status_val = result.get("status", "success")
                month_result["status"] = status_val
                if status_val != "success":
                    overall_status = "partial_success"
                total_written += result.get("written_patch_count", 0)
                total_skipped += result.get("skipped_duplicate_count", 0)
                total_failed += result.get("failed_patch_count", 0)
            except Exception as exc:
                month_result["status"] = "failed"
                month_result["message"] = str(exc)[:500]
                overall_status = "partial_success"

            per_month.append(month_result)

        return {
            "status": overall_status,
            "year": year,
            "mode": "all_months",
            "total_db_records": sum(m.get("db_records_count", 0) for m in per_month),
            "total_written_patch_count": total_written,
            "total_skipped_duplicate_count": total_skipped,
            "total_failed_patch_count": total_failed,
            "per_month_results": per_month,
        }

    def import_records(self, db: Session, operator_id: int | None = None, **kwargs: Any) -> SyncImportResult:
        """Import Tencent Docs matrix records through the shared ImportService."""

        year = int(kwargs.get("year") or get_tencent_docs_config()["default_year"])
        month = self._active_month(kwargs.get("month"))
        records = self.fetch_tencent_doc_rows(year=year, month=month)
        result = ImportService(db=db, operator_id=operator_id).import_records(records)
        parse_result = getattr(self, "_last_parse_result", SyncImportResult())
        result.failed += parse_result.failed
        result.errors = parse_result.errors + result.errors
        matrix_parse = getattr(self, "_last_matrix_parse", None)
        if matrix_parse is not None:
            result.extra_detail = {
                "detected_template_type": MATRIX_TEMPLATE_TYPE,
                "year": year,
                "month": month,
                "raw_rows_count": matrix_parse.raw_rows_count,
                "parsed_records_count": matrix_parse.parsed_records_count,
                "inserted_count": result.created,
                "skipped_count": result.skipped,
                "failed_count": result.failed,
                "invalid_records_preview": matrix_parse.invalid_records_preview(),
                "parsed_records_preview": matrix_parse.parsed_records_preview(),
                "reagent_count": matrix_parse.reagent_count,
                "reagent_names": matrix_parse.reagent_names,
            }
        return result

    def _export_matrix_current_month(self, db: Session, year: int = 2026, **kwargs: Any) -> dict[str, Any]:
        """Safely append local records by writing only changed 1x3 ranges."""

        self.validate_config()
        month = self._active_month(kwargs.get("month"))
        export_preview = self.preview_export_matrix(db=db, year=year, month=month)
        export_preview_response = {
            key: value
            for key, value in export_preview.items()
            if key not in ("values", "write_values")
        }

        if export_preview.get("db_records_count", 0) == 0:
            return {
                "success": False,
                "status": "no_data",
                "message": "当前年月数据库中没有可同步到腾讯文档的出入库流水，请先导入腾讯文档或新增出入库记录。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        if export_preview.get("patch_count", 0) == 0:
            return {
                "success": False,
                "status": "no_matched_records",
                "message": "当前年月数据库流水没有匹配到 19 种矩阵模板试剂列，未写入腾讯文档。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        if export_preview.get("failed_read_count", 0) > 0 and export_preview.get("will_write_patch_count", 0) == 0:
            return {
                "success": False,
                "status": "read_failed",
                "message": "无法读取腾讯文档已有相关单元格，为避免覆盖原表，已取消同步。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        if export_preview.get("will_write_patch_count", 0) == 0:
            return {
                "success": False,
                "status": "no_new_data",
                "message": "数据库记录已存在于腾讯文档，无需重复同步。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        written_ranges: list[str] = []
        failed_patches = list(export_preview.get("failed_patches") or [])
        written_patch_count = 0
        for patch in export_preview.get("patches", []):
            if not patch.get("will_write"):
                continue
            patch_values, validation_errors = validate_patch_values(patch.get("values") or [])
            if validation_errors:
                failed_patches.append(
                    {
                        "day": patch.get("day"),
                        "reagent_name": patch.get("reagent_name"),
                        "range": patch.get("full_range"),
                        "reason": "patch payload 校验失败",
                        "validation_errors": validation_errors,
                    }
                )
                continue
            try:
                self.write_sheet_range(patch["full_range"].split("!", 1)[0], patch["range"], patch_values)
            except TencentDocsApiError as exc:
                failed_patches.append(
                    {
                        "day": patch.get("day"),
                        "reagent_name": patch.get("reagent_name"),
                        "range": patch.get("full_range"),
                        "reason": "腾讯文档小范围写入失败",
                        "error": exc.detail or {"error": str(exc)},
                    }
                )
                continue
            written_ranges.append(patch["full_range"])
            written_patch_count += 1

        failed_patch_count = len(failed_patches)
        status_value = "success" if failed_patch_count == 0 else "partial_success"
        return {
            "success": failed_patch_count == 0,
            "status": status_value,
            "message": (
                f"腾讯文档追加同步完成：读取并检查 {export_preview.get('patch_count', 0)} 个三联单元格，"
                f"写入 {written_patch_count} 个小范围，"
                f"跳过重复 {export_preview.get('skipped_duplicate_count', 0)} 个，"
                f"失败 {failed_patch_count} 个，没有清空原表。"
            ),
            "year": year,
            "month": month,
            "sheet_id": export_preview.get("sheet_id"),
            "sheet_title": export_preview.get("sheet_title"),
            "write_range": "incremental_patches",
            "written_sheets": 1 if written_patch_count else 0,
            "total_rows": written_patch_count,
            "db_records_count": export_preview.get("db_records_count", 0),
            "patch_count": export_preview.get("patch_count", 0),
            "written_patch_count": written_patch_count,
            "will_write_patch_count": export_preview.get("will_write_patch_count", 0),
            "failed_patch_count": failed_patch_count,
            "failed_read_count": export_preview.get("failed_read_count", 0),
            "written_ranges": written_ranges,
            "failed_patches": failed_patches,
            "existing_cells_count": export_preview.get("existing_cells_count", 0),
            "new_tuples_count": export_preview.get("new_tuples_count", 0),
            "appended_tuples_count": export_preview.get("appended_tuples_count", 0),
            "skipped_duplicate_count": export_preview.get("skipped_duplicate_count", 0),
            "skipped_duplicate_tuples_count": export_preview.get("skipped_duplicate_tuples_count", 0),
            "changed_cells_count": export_preview.get("changed_cells_count", 0),
            "preview_changes": export_preview.get("preview_changes", []),
            "unmatched_records": export_preview.get("unmatched_records", []),
        }

    def _export_matrix_current_month(self, db: Session, year: int = 2026, **kwargs: Any) -> dict[str, Any]:
        """Append local records by writing only changed day+reagent 1x3 ranges.

        This is intentionally patch-based: it never writes the whole B4:BF34
        block. Each patch is read first, merged with local tuples, deduplicated,
        and then written back only when that specific 1x3 range changes.
        """

        self.validate_config()
        month = self._active_month(kwargs.get("month"))
        export_preview = self.preview_export_matrix(db=db, year=year, month=month)
        export_preview_response = {
            key: value
            for key, value in export_preview.items()
            if key not in ("values", "write_values")
        }

        if export_preview.get("db_records_count", 0) == 0:
            return {
                "success": False,
                "status": "no_data",
                "message": "当前年月数据库中没有可同步到腾讯文档的出入库流水，请先导入腾讯文档或新增出入库记录。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        if export_preview.get("patch_count", 0) == 0:
            return {
                "success": False,
                "status": "no_matched_records",
                "message": "当前年月数据库流水没有匹配到 19 种矩阵模板试剂列，未写入腾讯文档。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        if export_preview.get("will_write_patch_count", 0) == 0:
            return {
                "success": False,
                "status": "no_new_data",
                "message": "数据库记录已存在于腾讯文档，无需重复同步。",
                **export_preview_response,
                "written_sheets": 0,
                "total_rows": 0,
            }

        sheet_id = export_preview.get("sheet_id")
        written_ranges: list[str] = []
        failed_patches = list(export_preview.get("failed_patches") or [])
        written_patch_count = 0

        for patch in export_preview.get("patches", []):
            if not patch.get("will_write"):
                continue
            patch_values, validation_errors = validate_patch_values(patch.get("values") or [])
            if validation_errors:
                failed_patches.append(
                    {
                        "day": patch.get("day"),
                        "reagent_name": patch.get("reagent_name"),
                        "range": patch.get("full_range"),
                        "reason": "patch payload 校验失败",
                        "validation_errors": validation_errors,
                    }
                )
                continue
            try:
                self.write_sheet_range(sheet_id or "", patch["range"], patch_values)
            except TencentDocsApiError as exc:
                failed_patches.append(
                    {
                        "day": patch.get("day"),
                        "reagent_name": patch.get("reagent_name"),
                        "range": patch.get("full_range"),
                        "reason": "腾讯文档小范围写入失败",
                        "error": exc.detail or {"error": str(exc)},
                    }
                )
                continue
            written_ranges.append(patch.get("full_range") or f"{sheet_id}!{patch['range']}")
            written_patch_count += 1

        failed_patch_count = len(failed_patches)
        status_value = "success" if failed_patch_count == 0 else "partial_success"
        return {
            "success": failed_patch_count == 0,
            "status": status_value,
            "message": (
                f"腾讯文档追加同步完成：读取并检查 {export_preview.get('patch_count', 0)} 个三联单元格，"
                f"写入 {written_patch_count} 个小范围，"
                f"跳过重复 {export_preview.get('skipped_duplicate_count', 0)} 个，"
                f"失败 {failed_patch_count} 个，没有清空原表。"
            ),
            "year": year,
            "month": month,
            "sheet_id": sheet_id,
            "sheet_title": export_preview.get("sheet_title"),
            "write_range": "incremental_patches",
            "data_area_range": export_preview.get("data_area_range"),
            "written_sheets": 1 if written_patch_count else 0,
            "total_rows": written_patch_count,
            "db_records_count": export_preview.get("db_records_count", 0),
            "patch_count": export_preview.get("patch_count", 0),
            "written_patch_count": written_patch_count,
            "will_write_patch_count": export_preview.get("will_write_patch_count", 0),
            "failed_patch_count": failed_patch_count,
            "failed_read_count": export_preview.get("failed_read_count", 0),
            "written_ranges": written_ranges,
            "failed_patches": failed_patches,
            "existing_cells_count": export_preview.get("existing_cells_count", 0),
            "new_tuples_count": export_preview.get("new_tuples_count", 0),
            "appended_tuples_count": export_preview.get("appended_tuples_count", 0),
            "skipped_duplicate_count": export_preview.get("skipped_duplicate_count", 0),
            "skipped_duplicate_tuples_count": export_preview.get("skipped_duplicate_tuples_count", 0),
            "changed_cells_count": export_preview.get("changed_cells_count", 0),
            "preview_changes": export_preview.get("preview_changes", []),
            "unmatched_records": export_preview.get("unmatched_records", []),
        }

    def export_records(self, db: Session, year: int = 2026, **kwargs: Any) -> Any:
        """Write the selected month back to Tencent Docs with incremental patches.

        The formal export never writes the whole B4:BF34 block. It only writes
        changed 1x3 operation/quantity/operator ranges such as H28:J28.
        """

        return self._export_matrix_current_month(db=db, year=year, **kwargs)


def log_import_result(
    db: Session,
    provider_source: str,
    sync_type: str,
    result: SyncImportResult,
    message_prefix: str,
) -> SyncLog:
    """统一记录导入结果日志。"""

    message = (
        f"{message_prefix}完成，新增 {result.created} 条，"
        f"跳过 {result.skipped} 条，失败 {result.failed} 条"
    )
    return SyncLogService(db).create_log(
        source=provider_source,
        sync_type=sync_type,
        status_value="success" if result.failed == 0 else "partial_success",
        message=message,
        detail=result.to_detail_json(),
    )


def log_export_result(
    db: Session,
    provider_source: str,
    sync_type: str,
    message: str,
    detail: dict[str, Any] | None = None,
    status_value: str = "success",
) -> SyncLog:
    """统一记录导出结果日志。"""

    return SyncLogService(db).create_log(
        source=provider_source,
        sync_type=sync_type,
        status_value=status_value,
        message=message,
        detail=detail,
    )
